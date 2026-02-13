# ============================================================================
# BIBLIOTHÈQUE MEUBLES PARAMÉTRIQUES v5 - FreeCAD 1.0.2
# ============================================================================
# Gestion avancée du fond :
#   - Assemblage par vissage (entre panneaux)
#   - Assemblage par rainure/embrèvement (dans les 4 panneaux)
#   - Décalage arrière paramétrable
#   - Hauteur imposée optionnelle
# ============================================================================

import FreeCAD as App
import Part
from enum import Enum
from dataclasses import dataclass, field
from typing import List, Optional, Dict, Tuple, Any

# ============================================================================
# ÉNUMÉRATIONS
# ============================================================================

class Position(Enum):
    GAUCHE = "gauche"
    DROITE = "droite"
    HAUT = "haut"
    BAS = "bas"
    ARRIERE = "arriere"
    AVANT = "avant"
    CENTRE = "centre"


class TypeOuverture(Enum):
    GAUCHE = "gauche"
    DROITE = "droite"
    HAUT = "haut"
    BAS = "bas"


class TypeAssemblage(Enum):
    DESSUS_ENTRE = "dessus_entre"
    DESSUS_SUR = "dessus_sur"
    ABOUT = "about"


class TypePose(Enum):
    """Type de pose des charnières / portes.
    
    EN_APPLIQUE : recouvrement total (16mm) — la porte recouvre le panneau latéral
    SEMI_APPLIQUE : recouvrement partiel (8mm) — la porte recouvre la moitié du panneau
    ENCLOISONNEE : la porte est entre les panneaux, face extérieure alignée avec le corps,
                   jeu de 4mm entre la porte et le panneau
    """
    EN_APPLIQUE = "applique"
    SEMI_APPLIQUE = "semi_applique"
    ENCLOISONNEE = "encloisonnee"


class TypeAssemblageFond(Enum):
    """Type d'assemblage du fond"""
    VISSAGE = "vissage"           # Vissé entre les panneaux (le plus simple)
    RAINURE = "rainure"           # Embrèvement dans les 4 panneaux
    APPLIQUE = "applique"         # Appliqué sur l'arrière (vissé sur chants)


# ============================================================================
# CONFIGURATION DU FOND
# ============================================================================

@dataclass
class ConfigFond:
    """
    Configuration complète du panneau de fond.
    
    Le fond peut être assemblé de plusieurs façons :
    
    1. VISSAGE : Le fond est vissé entre les panneaux
       - Largeur = largeur_intérieure (entre côtés)
       - Hauteur = hauteur entre dessus et dessous
       - Position = décalé depuis l'arrière
    
    2. RAINURE : Le fond est embrèvé dans une rainure
       - Largeur = largeur_intérieure + 2 * profondeur_rainure
       - Hauteur = hauteur + 2 * profondeur_rainure
       - Rainure usinée dans les 4 panneaux
    
    3. APPLIQUE : Le fond est vissé sur les chants arrière
       - Largeur = largeur totale du meuble
       - Hauteur = hauteur totale du corps
    """
    
    # Type d'assemblage
    type_assemblage: TypeAssemblageFond = TypeAssemblageFond.RAINURE
    
    # Épaisseur du fond (3mm HDF, 5mm CP, 8mm, 10mm, 16mm, 19mm)
    epaisseur: float = 3.0
    
    # Décalage depuis l'arrière du meuble (0 = au ras de l'arrière)
    decalage_arriere: float = 0.0
    
    # --- Paramètres pour RAINURE ---
    profondeur_rainure: float = 8.0     # Profondeur de la rainure dans les panneaux
    jeu_fond_rainure: float = 1.0       # Jeu du fond dans la rainure (pour dilatation)
    distance_rainure_chant: float = 10.0  # Distance de la rainure depuis le chant arrière
    
    # --- Paramètres pour VISSAGE ---
    # Le fond est positionné entre les panneaux
    # Pas de paramètres supplémentaires
    
    # --- Hauteur imposée (optionnel) ---
    hauteur_imposee: Optional[float] = None  # Si défini, limite la hauteur du fond
    position_verticale: str = "bas"          # "bas", "haut", "centre" - où placer le fond si hauteur imposée
    
    # --- Matériau ---
    couleur: Tuple[float, float, float] = (0.85, 0.75, 0.60)
    
    def avec_vissage(self) -> 'ConfigFond':
        """Configure pour assemblage par vissage"""
        self.type_assemblage = TypeAssemblageFond.VISSAGE
        return self
    
    def avec_rainure(
        self, 
        profondeur: float = 8.0, 
        jeu: float = 1.0,
        distance_chant: float = 10.0
    ) -> 'ConfigFond':
        """Configure pour assemblage par rainure/embrèvement"""
        self.type_assemblage = TypeAssemblageFond.RAINURE
        self.profondeur_rainure = profondeur
        self.jeu_fond_rainure = jeu
        self.distance_rainure_chant = distance_chant
        return self
    
    def avec_applique(self) -> 'ConfigFond':
        """Configure pour assemblage appliqué"""
        self.type_assemblage = TypeAssemblageFond.APPLIQUE
        return self
    
    def avec_decalage(self, decalage: float) -> 'ConfigFond':
        """Définit le décalage depuis l'arrière"""
        self.decalage_arriere = decalage
        return self
    
    def avec_hauteur_imposee(
        self, 
        hauteur: float, 
        position: str = "bas"
    ) -> 'ConfigFond':
        """
        Impose une hauteur maximale au fond.
        
        Args:
            hauteur: Hauteur maximale du fond
            position: "bas" (depuis le dessous), "haut" (depuis le dessus), "centre"
        """
        self.hauteur_imposee = hauteur
        self.position_verticale = position
        return self
    
    def sans_hauteur_imposee(self) -> 'ConfigFond':
        """Retire la limitation de hauteur"""
        self.hauteur_imposee = None
        return self
    
    @classmethod
    def standard_rainure(cls, epaisseur: float = 3.0) -> 'ConfigFond':
        """Configuration standard avec rainure"""
        return cls(
            type_assemblage=TypeAssemblageFond.RAINURE,
            epaisseur=epaisseur,
            decalage_arriere=0,
            profondeur_rainure=8.0,
            jeu_fond_rainure=1.0,
            distance_rainure_chant=10.0
        )
    
    @classmethod
    def standard_vissage(cls, epaisseur: float = 5.0) -> 'ConfigFond':
        """Configuration standard avec vissage"""
        return cls(
            type_assemblage=TypeAssemblageFond.VISSAGE,
            epaisseur=epaisseur,
            decalage_arriere=5.0  # Légèrement en retrait pour accès vissage
        )
    
    @classmethod
    def applique_arriere(cls, epaisseur: float = 3.0) -> 'ConfigFond':
        """Configuration appliqué à l'arrière"""
        return cls(
            type_assemblage=TypeAssemblageFond.APPLIQUE,
            epaisseur=epaisseur,
            decalage_arriere=0
        )


# ============================================================================
# CONFIGURATION DES CRÉMAILLÈRES
# ============================================================================

@dataclass
class ConfigCremaillere:
    """
    Configuration des crémaillères pour étagères réglables.
    
    Les crémaillères sont des profilés aluminium encastrés dans les côtés
    et séparations verticales du meuble.
    
    Paramètres standard crémaillère aluminium:
    - Largeur: 16mm
    - Profondeur (encastrement): 7mm
    - Pas des trous: 32mm (système 32)
    """
    
    # Dimensions de la crémaillère
    largeur: float = 16.0               # Largeur du profilé
    profondeur: float = 7.0             # Profondeur d'encastrement (rainure)
    
    # Positionnement
    distance_avant: float = 37.0        # Distance depuis l'avant du panneau
    distance_arriere: float = 37.0      # Distance depuis l'arrière du panneau
    
    # Si True, 2 crémaillères par panneau (avant + arrière)
    # Si False, 1 seule crémaillère (positionnée selon position_unique)
    double: bool = True
    position_unique: str = "avant"      # "avant" ou "arriere" si double=False
    
    # Couleur pour visualisation
    couleur: Tuple[float, float, float] = (0.75, 0.75, 0.78)  # Aluminium
    
    def avec_positions(self, avant: float, arriere: float) -> 'ConfigCremaillere':
        """Définit les distances avant et arrière"""
        self.distance_avant = avant
        self.distance_arriere = arriere
        return self
    
    def avec_simple(self, position: str = "avant") -> 'ConfigCremaillere':
        """Configure une seule crémaillère par panneau"""
        self.double = False
        self.position_unique = position
        return self
    
    @classmethod
    def standard(cls) -> 'ConfigCremaillere':
        """Configuration standard avec 2 crémaillères"""
        return cls(
            largeur=16.0,
            profondeur=7.0,
            distance_avant=37.0,
            distance_arriere=37.0,
            double=True
        )
    
    @classmethod
    def simple_avant(cls) -> 'ConfigCremaillere':
        """Une seule crémaillère à l'avant"""
        return cls(
            largeur=16.0,
            profondeur=7.0,
            distance_avant=37.0,
            double=False,
            position_unique="avant"
        )


# ============================================================================
# CONFIGURATION DES ÉTAGÈRES
# ============================================================================

@dataclass
class ConfigEtagere:
    """
    Configuration des étagères sur crémaillères.
    
    Les étagères sont posées sur des taquets insérés dans les crémaillères.
    Elles ont des jeux par rapport aux panneaux latéraux pour faciliter
    la mise en place.
    """
    
    # Jeux
    jeu_lateral: float = 1.0            # Jeu de chaque côté (entre étagère et côté/séparation)
    jeu_arriere: float = 5.0            # Jeu avec le fond
    retrait_avant: float = 20.0         # Retrait par rapport à l'avant du corps
    
    # Chants
    chant_avant: bool = True            # Chant sur la face avant
    chant_arriere: bool = False         # Chant sur la face arrière
    chant_lateraux: bool = False        # Chants sur les côtés
    epaisseur_chant: float = 1.0
    
    def avec_jeux(self, lateral: float = 1.0, arriere: float = 5.0) -> 'ConfigEtagere':
        """Définit les jeux"""
        self.jeu_lateral = lateral
        self.jeu_arriere = arriere
        return self
    
    def avec_retrait_avant(self, retrait: float) -> 'ConfigEtagere':
        """Définit le retrait avant"""
        self.retrait_avant = retrait
        return self
    
    def avec_chants(self, avant: bool = True, arriere: bool = False, lateraux: bool = False) -> 'ConfigEtagere':
        """Configure les chants"""
        self.chant_avant = avant
        self.chant_arriere = arriere
        self.chant_lateraux = lateraux
        return self
    
    @classmethod
    def standard(cls) -> 'ConfigEtagere':
        """Configuration standard"""
        return cls(
            jeu_lateral=1.0,
            jeu_arriere=5.0,
            retrait_avant=20.0,
            chant_avant=True
        )


# ============================================================================
# CONFIGURATION DES PIEDS ET PLINTHES
# ============================================================================

@dataclass
class ConfigPlinthe:
    """
    Configuration des pieds réglables et des plinthes.
    
    Le meuble est posé sur des pieds réglables en hauteur.
    Les plinthes se clipsent ou se fixent sur les pieds et peuvent
    être présentes sur 1, 2 ou 3 côtés du meuble.
    
    Côtés disponibles:
    - "avant"  : face avant (toujours présente par défaut)
    - "gauche" : côté gauche
    - "droite" : côté droit
    
    Note: pas de plinthe à l'arrière (le meuble est contre le mur).
    """
    
    # Dimensions
    hauteur: float = 100.0          # Hauteur de la plinthe (= hauteur_plinthe du meuble)
    epaisseur: float = 16.0         # Épaisseur de la plinthe
    
    # Retraits par rapport aux faces du meuble
    retrait: float = 30.0           # Retrait depuis la face avant (façade)
    retrait_lateral: float = 16.0   # Retrait des plinthes latérales depuis le côté du meuble
    
    # Côtés avec plinthe
    cotes: List[str] = field(default_factory=lambda: ["avant"])
    
    # Pieds
    diametre_pied: float = 30.0     # Diamètre du pied (pour visualisation)
    hauteur_pied: float = 100.0     # Hauteur du pied (sera = hauteur plinthe)
    nombre_pieds_profondeur: int = 2   # Nombre de pieds sur la profondeur (avant/arrière)
    nombre_pieds_largeur: int = 2      # Nombre de pieds sur la largeur (gauche/droite)
    marge_pied: float = 50.0        # Distance minimale du pied au bord du meuble
    
    # Matériau / couleur
    couleur_plinthe: Tuple[float, float, float] = (0.25, 0.25, 0.25)  # Gris foncé
    couleur_pied: Tuple[float, float, float] = (0.3, 0.3, 0.3)       # Gris
    
    def avec_cotes(self, cotes: List[str]) -> 'ConfigPlinthe':
        """Définit les côtés avec plinthe"""
        valides = {"avant", "gauche", "droite"}
        self.cotes = [c for c in cotes if c in valides]
        return self
    
    def avec_retrait(self, retrait: float, lateral: float = None) -> 'ConfigPlinthe':
        """Définit le retrait des plinthes (avant et optionnellement latéral)"""
        self.retrait = retrait
        if lateral is not None:
            self.retrait_lateral = lateral
        return self
    
    def avec_pieds(self, profondeur: int = 2, largeur: int = 2) -> 'ConfigPlinthe':
        """Définit le nombre de pieds"""
        self.nombre_pieds_profondeur = profondeur
        self.nombre_pieds_largeur = largeur
        return self
    
    @classmethod
    def avant_seul(cls, hauteur: float = 100.0) -> 'ConfigPlinthe':
        """Plinthe sur la face avant uniquement"""
        return cls(hauteur=hauteur, cotes=["avant"])
    
    @classmethod
    def trois_cotes(cls, hauteur: float = 100.0) -> 'ConfigPlinthe':
        """Plinthe sur avant, gauche et droite"""
        return cls(hauteur=hauteur, cotes=["avant", "gauche", "droite"])
    
    @classmethod
    def en_l_gauche(cls, hauteur: float = 100.0) -> 'ConfigPlinthe':
        """Plinthe sur avant et gauche (meuble en bout à droite)"""
        return cls(hauteur=hauteur, cotes=["avant", "gauche"])
    
    @classmethod
    def en_l_droite(cls, hauteur: float = 100.0) -> 'ConfigPlinthe':
        """Plinthe sur avant et droite (meuble en bout à gauche)"""
        return cls(hauteur=hauteur, cotes=["avant", "droite"])


# ============================================================================
# QUINCAILLERIE BLUM
# ============================================================================

class LegraboxHauteur(Enum):
    M = "M"      # 90.5 mm
    K = "K"      # 128.5 mm
    C = "C"      # 193 mm
    F = "F"      # 257 mm


@dataclass
class LegraboxSpec:
    jeu_lateral: float = 12.75
    jeu_total_lateral: float = 25.5
    epaisseur_paroi: float = 12.5
    epaisseur_fond_standard: float = 8
    
    hauteurs = {
        LegraboxHauteur.M: 90.5,
        LegraboxHauteur.K: 128.5,
        LegraboxHauteur.C: 193.0,
        LegraboxHauteur.F: 257.0,
    }
    longueurs = [270, 300, 350, 400, 450, 500, 550, 600, 650]
    
    @classmethod
    def get_hauteur_cote(cls, hauteur: LegraboxHauteur) -> float:
        return cls.hauteurs.get(hauteur, 90.5)
    
    @classmethod
    def calculer_profondeur_tiroir(cls, profondeur_caisson: float, ep_facade: float = 19) -> int:
        prof_dispo = profondeur_caisson - ep_facade - 10
        for longueur in sorted(cls.longueurs, reverse=True):
            if longueur <= prof_dispo:
                return longueur
        return cls.longueurs[0]


@dataclass
class ClipTopSpec:
    diametre_cuvette: float = 35.0
    profondeur_cuvette: float = 13.0
    # Distance du bord de la cuvette au bord de la porte (côté charnière)
    distance_bord_cuvette: float = 3.0
    # Recouvrements
    recouvrement_applique: float = 16.0       # Recouvrement en applique
    recouvrement_semi_applique: float = 8.0   # Recouvrement en semi-applique
    jeu_encloisonnee: float = 4.0             # Jeu porte/panneau en encloisonné
    
    @classmethod
    def nombre_charnieres(cls, hauteur_porte: float) -> int:
        if hauteur_porte <= 1000:
            return 2
        elif hauteur_porte <= 1500:
            return 3
        elif hauteur_porte <= 2000:
            return 4
        return 5
    
    @classmethod
    def positions_charnieres(cls, hauteur_porte: float) -> List[float]:
        nb = cls.nombre_charnieres(hauteur_porte)
        distance_bord = 80.0
        
        if nb == 2:
            return [distance_bord, hauteur_porte - distance_bord]
        
        espace_dispo = hauteur_porte - 2 * distance_bord
        espacement = espace_dispo / (nb - 1)
        return [distance_bord + i * espacement for i in range(nb)]
    
    @classmethod
    def centre_cuvette_depuis_bord(cls) -> float:
        """Distance du centre de la cuvette au bord de la porte (côté charnière)"""
        return cls.distance_bord_cuvette + cls.diametre_cuvette / 2


# ============================================================================
# CONFIGURATION DES SÉPARATIONS
# ============================================================================

@dataclass
class ConfigSeparations:
    """
    Configuration des séparations verticales du meuble.
    
    Deux modes de définition:
    1. Par nombre de compartiments égaux
    2. Par liste des largeurs de compartiments
    """
    
    # Mode 1: Compartiments égaux
    nombre_compartiments: int = 1           # 1 = pas de séparation
    
    # Mode 2: Largeurs personnalisées (prioritaire si défini)
    largeurs_compartiments: Optional[List[float]] = None
    
    # Épaisseur des séparations
    epaisseur: float = 19
    
    # Retraits
    retrait_arriere: float = 0              # Distance depuis le fond
    retrait_avant: float = 0                # Distance depuis l'avant du corps
    
    # Chants
    chant_avant: bool = True
    epaisseur_chant: float = 1
    
    def avec_compartiments_egaux(self, nombre: int) -> 'ConfigSeparations':
        """Définit un nombre de compartiments égaux"""
        self.nombre_compartiments = nombre
        self.largeurs_compartiments = None
        return self
    
    def avec_largeurs(self, largeurs: List[float]) -> 'ConfigSeparations':
        """Définit les largeurs personnalisées des compartiments"""
        self.largeurs_compartiments = largeurs
        self.nombre_compartiments = len(largeurs)
        return self
    
    def avec_retraits(self, arriere: float = 0, avant: float = 0) -> 'ConfigSeparations':
        """Définit les retraits avant/arrière"""
        self.retrait_arriere = arriere
        self.retrait_avant = avant
        return self
    
    @classmethod
    def sans_separation(cls) -> 'ConfigSeparations':
        """Pas de séparation (1 seul compartiment)"""
        return cls(nombre_compartiments=1)
    
    @classmethod
    def deux_compartiments(cls, epaisseur: float = 19) -> 'ConfigSeparations':
        """2 compartiments égaux"""
        return cls(nombre_compartiments=2, epaisseur=epaisseur)
    
    @classmethod
    def trois_compartiments(cls, epaisseur: float = 19) -> 'ConfigSeparations':
        """3 compartiments égaux"""
        return cls(nombre_compartiments=3, epaisseur=epaisseur)


# ============================================================================
# RÈGLES D'ASSEMBLAGE
# ============================================================================

@dataclass
class ReglesAssemblage:
    type_assemblage: TypeAssemblage = TypeAssemblage.DESSUS_ENTRE
    type_pose: TypePose = TypePose.EN_APPLIQUE
    
    # Jeux façades
    jeu_porte_haut: float = 4
    jeu_porte_bas: float = 4
    jeu_porte_lateral: float = 2
    jeu_entre_portes: float = 3
    jeu_tiroir_lateral: float = 2
    jeu_entre_tiroirs: float = 4
    
    # Jeux intérieurs (obsolète - utiliser config_etagere)
    jeu_etagere_lateral: float = 1
    jeu_etagere_profondeur: float = 20
    
    # Configuration du fond
    config_fond: ConfigFond = field(default_factory=ConfigFond.standard_rainure)
    
    # Configuration des séparations
    config_separations: ConfigSeparations = field(default_factory=ConfigSeparations.sans_separation)
    
    # Configuration des crémaillères
    config_cremaillere: ConfigCremaillere = field(default_factory=ConfigCremaillere.standard)
    
    # Configuration des étagères
    config_etagere: ConfigEtagere = field(default_factory=ConfigEtagere.standard)
    
    # Configuration des pieds et plinthes
    config_plinthe: ConfigPlinthe = field(default_factory=ConfigPlinthe.avant_seul)
    
    # Quincaillerie
    legrabox_hauteur: LegraboxHauteur = LegraboxHauteur.M
    
    def configurer_pose(self, type_pose: TypePose) -> 'ReglesAssemblage':
        """Configure le type de pose des portes/charnières"""
        self.type_pose = type_pose
        return self
    
    def configurer_fond(self, config: ConfigFond) -> 'ReglesAssemblage':
        """Configure le fond du meuble"""
        self.config_fond = config
        return self
    
    def configurer_separations(self, config: ConfigSeparations) -> 'ReglesAssemblage':
        """Configure les séparations verticales"""
        self.config_separations = config
        return self
    
    def configurer_cremailleres(self, config: ConfigCremaillere) -> 'ReglesAssemblage':
        """Configure les crémaillères"""
        self.config_cremaillere = config
        return self
    
    def configurer_etageres(self, config: ConfigEtagere) -> 'ReglesAssemblage':
        """Configure les étagères"""
        self.config_etagere = config
        return self
    
    def configurer_plinthe(self, config: ConfigPlinthe) -> 'ReglesAssemblage':
        """Configure les pieds et plinthes"""
        self.config_plinthe = config
        return self
    
    @classmethod
    def standard(cls) -> 'ReglesAssemblage':
        return cls(
            config_fond=ConfigFond.standard_rainure(),
            config_separations=ConfigSeparations.sans_separation(),
            config_cremaillere=ConfigCremaillere.standard(),
            config_etagere=ConfigEtagere.standard(),
            config_plinthe=ConfigPlinthe.avant_seul()
        )
    
    @classmethod
    def standard_vissage(cls) -> 'ReglesAssemblage':
        return cls(
            config_fond=ConfigFond.standard_vissage(),
            config_separations=ConfigSeparations.sans_separation(),
            config_cremaillere=ConfigCremaillere.standard(),
            config_etagere=ConfigEtagere.standard(),
            config_plinthe=ConfigPlinthe.avant_seul()
        )


# ============================================================================
# MATÉRIAUX
# ============================================================================

@dataclass
class Materiau:
    nom: str
    epaisseur: float
    couleur: Tuple[float, float, float] = (0.82, 0.71, 0.55)
    
    @classmethod
    def melamine_blanc(cls):
        return cls("Mélaminé Blanc", 19, (0.95, 0.95, 0.95))
    
    @classmethod
    def melamine_chene(cls):
        return cls("Mélaminé Chêne", 19, (0.82, 0.71, 0.55))
    
    @classmethod
    def fond_hdf_3(cls):
        return cls("HDF 3mm", 3, (0.85, 0.75, 0.60))
    
    @classmethod
    def fond_cp_5(cls):
        return cls("CP 5mm", 5, (0.80, 0.70, 0.55))
    
    @classmethod
    def fond_mdf_8(cls):
        return cls("MDF 8mm", 8, (0.76, 0.65, 0.50))


# ============================================================================
# GESTIONNAIRE D'OBJETS
# ============================================================================

class GestionnaireObjets:
    def __init__(self, doc):
        self.doc = doc
        self._objets_utilises = set()
    
    def get_ou_creer_objet(self, nom: str, type_objet: str = "Part::Feature") -> Any:
        obj = self.doc.getObject(nom)
        if obj is None:
            obj = self.doc.addObject(type_objet, nom)
        self._objets_utilises.add(nom)
        return obj
    
    def get_ou_creer_groupe(self, nom: str, parent=None) -> Any:
        obj = self.doc.getObject(nom)
        if obj is None:
            obj = self.doc.addObject("App::Part", nom)
            obj.Label = nom.replace("_", " ")
            if parent and obj not in parent.Group:
                parent.addObject(obj)
        self._objets_utilises.add(nom)
        return obj
    
    def mettre_a_jour_shape(self, obj, shape: Part.Shape):
        obj.Shape = shape
    
    def appliquer_couleur(self, obj, couleur: Tuple[float, float, float], transparence: int = 0):
        try:
            if hasattr(obj, "ViewObject") and obj.ViewObject:
                obj.ViewObject.ShapeColor = couleur
                if transparence > 0:
                    obj.ViewObject.Transparency = transparence
        except:
            pass
    
    def nettoyer_objets_orphelins(self, prefixe: str):
        a_supprimer = []
        for obj in self.doc.Objects:
            if obj.Name.startswith(prefixe) and obj.Name not in self._objets_utilises:
                a_supprimer.append(obj.Name)
        for nom in a_supprimer:
            try:
                self.doc.removeObject(nom)
            except:
                pass


# ============================================================================
# PANNEAU
# ============================================================================

@dataclass
class Panneau:
    nom: str
    longueur: float
    largeur: float
    epaisseur: float = 19
    position: Tuple[float, float, float] = (0, 0, 0)
    orientation: str = "XY"
    chants: Dict[str, float] = field(default_factory=dict)
    materiau: Optional[Materiau] = None
    couleur: Tuple[float, float, float] = (0.82, 0.71, 0.55)
    couleur_chant: Tuple[float, float, float] = (0.4, 0.4, 0.4)
    
    # Rainures pour fond
    rainures: List[Dict] = field(default_factory=list)
    
    # Rainures pour crémaillères
    rainures_cremaillere: List[Dict] = field(default_factory=list)
    
    def ajouter_chant(self, face: str, epaisseur: float = 1) -> 'Panneau':
        self.chants[face] = epaisseur
        return self
    
    def chants_tous(self, epaisseur: float = 1) -> 'Panneau':
        for face in ['avant', 'arriere', 'gauche', 'droite']:
            self.chants[face] = epaisseur
        return self
    
    def chants_visibles(self, epaisseur: float = 1) -> 'Panneau':
        self.chants['avant'] = epaisseur
        return self
    
    def ajouter_rainure(
        self, 
        position: float,       # Distance depuis le chant de référence
        largeur: float,        # Largeur de la rainure
        profondeur: float,     # Profondeur de la rainure
        direction: str = "Y"   # Direction de la rainure
    ) -> 'Panneau':
        """Ajoute une rainure pour le fond"""
        self.rainures.append({
            'position': position,
            'largeur': largeur,
            'profondeur': profondeur,
            'direction': direction
        })
        return self
    
    def ajouter_rainure_cremaillere(
        self,
        distance_avant: float,      # Distance depuis l'avant du panneau
        largeur: float = 16.0,      # Largeur de la rainure
        profondeur: float = 7.0,    # Profondeur de la rainure
        hauteur_debut: float = 0,   # Début de la rainure (depuis le bas)
        hauteur_fin: float = None   # Fin de la rainure (None = toute la hauteur)
    ) -> 'Panneau':
        """
        Ajoute une rainure pour crémaillère.
        La rainure est verticale, sur la face intérieure du panneau.
        """
        self.rainures_cremaillere.append({
            'distance_avant': distance_avant,
            'largeur': largeur,
            'profondeur': profondeur,
            'hauteur_debut': hauteur_debut,
            'hauteur_fin': hauteur_fin
        })
        return self
    
    def get_surface(self) -> float:
        return (self.longueur * self.largeur) / 1_000_000
    
    def get_lineaire_chant(self) -> Dict[str, float]:
        result = {}
        for face in self.chants.keys():
            if face in ['avant', 'arriere']:
                result[face] = self.largeur / 1000
            else:
                result[face] = self.longueur / 1000
        return result
    
    def construire(self, gestionnaire: GestionnaireObjets, prefixe: str = "") -> List:
        objets = []
        nom_complet = f"{prefixe}_{self.nom}" if prefixe else self.nom
        couleur = self.materiau.couleur if self.materiau else self.couleur
        ep = self.materiau.epaisseur if self.materiau else self.epaisseur
        
        if self.orientation == "XY":
            dims = (self.longueur, self.largeur, ep)
        elif self.orientation == "XZ":
            dims = (self.longueur, ep, self.largeur)
        else:
            dims = (ep, self.longueur, self.largeur)
        
        # Créer le shape de base
        shape = Part.makeBox(dims[0], dims[1], dims[2])
        shape.translate(App.Vector(*self.position))
        
        # Soustraire les rainures de fond
        for rainure in self.rainures:
            rainure_shape = self._creer_shape_rainure(rainure, dims, ep)
            if rainure_shape:
                try:
                    shape = shape.cut(rainure_shape)
                except:
                    pass
        
        # Soustraire les rainures de crémaillères
        for rainure in self.rainures_cremaillere:
            rainure_shape = self._creer_shape_rainure_cremaillere(rainure, dims, ep)
            if rainure_shape:
                try:
                    shape = shape.cut(rainure_shape)
                except:
                    pass
        
        obj = gestionnaire.get_ou_creer_objet(nom_complet)
        gestionnaire.mettre_a_jour_shape(obj, shape)
        gestionnaire.appliquer_couleur(obj, couleur)
        objets.append(obj)
        
        # Chants
        for face, ep_chant in self.chants.items():
            chant_obj = self._construire_chant(gestionnaire, nom_complet, face, ep_chant, dims)
            if chant_obj:
                objets.append(chant_obj)
        
        return objets
    
    def _creer_shape_rainure_cremaillere(self, rainure: Dict, dims: Tuple, ep: float) -> Part.Shape:
        """
        Crée le shape de la rainure de crémaillère à soustraire.
        
        La rainure est verticale (direction Z), sur la face intérieure du panneau.
        Pour un côté: la rainure est positionnée par rapport à l'AVANT du panneau.
        """
        distance_avant = rainure['distance_avant']
        largeur = rainure['largeur']
        profondeur = rainure['profondeur']
        hauteur_debut = rainure.get('hauteur_debut', 0)
        hauteur_fin = rainure.get('hauteur_fin', None)
        
        try:
            if self.orientation == "XZ":
                # Panneau vertical (côté gauche/droit)
                # Le panneau a: longueur(profondeur) en X, épaisseur en Y, largeur(hauteur) en Z
                
                # Position X: depuis l'AVANT du panneau (X = longueur du panneau)
                # distance_avant est mesurée depuis la face avant
                x_rainure = self.position[0] + self.longueur - distance_avant - largeur
                
                # Position Y: sur la face intérieure
                if "Gauche" in self.nom:
                    # Côté gauche: face intérieure = côté Y+
                    y_rainure = self.position[1] + dims[1] - profondeur
                else:
                    # Côté droit: face intérieure = côté Y-
                    y_rainure = self.position[1]
                
                # Position Z et hauteur
                z_rainure = self.position[2] + hauteur_debut
                hauteur = (hauteur_fin if hauteur_fin else self.largeur) - hauteur_debut
                
                rainure_shape = Part.makeBox(largeur, profondeur, hauteur)
                rainure_shape.translate(App.Vector(x_rainure, y_rainure, z_rainure))
                
                return rainure_shape
                
            else:
                # Pour les séparations (orientation YZ)
                return None
                
        except Exception as e:
            print(f"Erreur création rainure crémaillère pour {self.nom}: {e}")
            return None
    
    def _creer_shape_rainure(self, rainure: Dict, dims: Tuple, ep: float) -> Part.Shape:
        """
        Crée le shape de la rainure à soustraire du panneau.
        
        La rainure est définie par:
        - position: distance depuis le bord arrière du panneau (X=0 du panneau)
        - largeur: largeur de la rainure
        - profondeur: profondeur de la rainure (dans l'épaisseur du panneau)
        - direction: "Y" ou "Z" selon l'orientation de la rainure
        - face: "interieure" (défaut) ou "exterieure" - sur quelle face faire la rainure
        """
        pos_rainure = rainure['position']
        larg_rainure = rainure['largeur']
        prof_rainure = rainure['profondeur']
        direction = rainure.get('direction', 'Z')
        face = rainure.get('face', 'interieure')
        
        try:
            if self.orientation == "XZ":
                # Panneau vertical (côté gauche/droit)
                # Le panneau a: longueur en X, épaisseur en Y, largeur(hauteur) en Z
                
                # Position X de la rainure (depuis l'origine du panneau)
                x_rainure = self.position[0] + pos_rainure
                
                # La rainure est sur la face intérieure du panneau
                if "Gauche" in self.nom:
                    # Côté gauche: face intérieure = côté Y+
                    y_rainure = self.position[1] + dims[1] - prof_rainure
                else:
                    # Côté droit: face intérieure = côté Y-
                    y_rainure = self.position[1]
                
                z_rainure = self.position[2]
                
                # La rainure traverse toute la hauteur du panneau
                rainure_shape = Part.makeBox(
                    larg_rainure,      # Largeur en X
                    prof_rainure,      # Profondeur en Y
                    dims[2]            # Hauteur totale en Z
                )
                rainure_shape.translate(App.Vector(x_rainure, y_rainure, z_rainure))
                
            elif self.orientation == "XY":
                # Panneau horizontal (dessus/dessous/traverse)
                # Le panneau a: longueur en X, largeur en Y, épaisseur en Z
                
                # Position X de la rainure
                x_rainure = self.position[0] + pos_rainure
                y_rainure = self.position[1]
                
                # Déterminer sur quelle face (dessus ou dessous du panneau)
                if "Dessous" in self.nom:
                    # Panneau du dessous: rainure sur la face supérieure (Z+)
                    z_rainure = self.position[2] + dims[2] - prof_rainure
                elif "Dessus" in self.nom:
                    # Panneau du dessus: rainure sur la face inférieure (Z-)
                    z_rainure = self.position[2]
                elif "Traverse_Arriere" in self.nom:
                    # Traverse arrière: rainure sur la face supérieure
                    z_rainure = self.position[2] + dims[2] - prof_rainure
                else:
                    # Par défaut: face supérieure
                    z_rainure = self.position[2] + dims[2] - prof_rainure
                
                # La rainure traverse toute la largeur du panneau
                rainure_shape = Part.makeBox(
                    larg_rainure,      # Largeur en X
                    dims[1],           # Traverse toute la largeur en Y
                    prof_rainure       # Profondeur en Z
                )
                rainure_shape.translate(App.Vector(x_rainure, y_rainure, z_rainure))
                
            else:
                # Orientation YZ - pas utilisé pour les rainures de fond standard
                return None
            
            return rainure_shape
            
        except Exception as e:
            print(f"Erreur création rainure pour {self.nom}: {e}")
            return None
    
    def _construire_chant(self, gestionnaire, nom_panneau, face, ep_chant, dims):
        pos = list(self.position)
        nom_chant = f"{nom_panneau}_chant_{face}"
        
        if self.orientation == "XY":
            if face == 'avant':
                chant_dims = (ep_chant, dims[1], dims[2])
                pos[0] += dims[0]
            elif face == 'arriere':
                chant_dims = (ep_chant, dims[1], dims[2])
                pos[0] -= ep_chant
            elif face == 'gauche':
                chant_dims = (dims[0], ep_chant, dims[2])
                pos[1] -= ep_chant
            elif face == 'droite':
                chant_dims = (dims[0], ep_chant, dims[2])
                pos[1] += dims[1]
            else:
                return None
        elif self.orientation == "XZ":
            if face == 'avant':
                chant_dims = (ep_chant, dims[1], dims[2])
                pos[0] += dims[0]
            elif face == 'arriere':
                chant_dims = (ep_chant, dims[1], dims[2])
                pos[0] -= ep_chant
            else:
                return None
        else:
            if face == 'avant':
                chant_dims = (dims[0], ep_chant, dims[2])
                pos[1] += dims[1]
            else:
                return None
        
        shape = Part.makeBox(*chant_dims)
        shape.translate(App.Vector(*pos))
        
        obj = gestionnaire.get_ou_creer_objet(nom_chant)
        gestionnaire.mettre_a_jour_shape(obj, shape)
        gestionnaire.appliquer_couleur(obj, self.couleur_chant)
        
        return obj


# ============================================================================
# PANNEAU DE FOND
# ============================================================================

@dataclass
class PanneauFond:
    """
    Panneau de fond avec gestion avancée de l'assemblage.
    """
    nom: str
    largeur: float
    hauteur: float
    epaisseur: float
    position: Tuple[float, float, float]
    config: ConfigFond
    couleur: Tuple[float, float, float] = (0.85, 0.75, 0.60)
    
    def get_surface(self) -> float:
        return (self.largeur * self.hauteur) / 1_000_000
    
    def construire(self, gestionnaire: GestionnaireObjets, prefixe: str = "") -> Any:
        nom_complet = f"{prefixe}_{self.nom}" if prefixe else self.nom
        
        # Le fond est orienté YZ (perpendiculaire à X)
        shape = Part.makeBox(self.epaisseur, self.largeur, self.hauteur)
        shape.translate(App.Vector(*self.position))
        
        obj = gestionnaire.get_ou_creer_objet(nom_complet)
        gestionnaire.mettre_a_jour_shape(obj, shape)
        gestionnaire.appliquer_couleur(obj, self.couleur)
        
        return obj
    
    def generer_info_usinage(self) -> str:
        """Génère les informations d'usinage pour le fond"""
        lines = [
            f"FOND - {self.nom}",
            f"  Dimensions: {self.largeur:.1f} x {self.hauteur:.1f} x {self.epaisseur}mm",
            f"  Surface: {self.get_surface():.4f} m²",
            f"  Type assemblage: {self.config.type_assemblage.value}",
        ]
        
        if self.config.type_assemblage == TypeAssemblageFond.RAINURE:
            lines.extend([
                f"  Rainure:",
                f"    - Profondeur: {self.config.profondeur_rainure}mm",
                f"    - Jeu: {self.config.jeu_fond_rainure}mm",
                f"    - Distance chant arrière: {self.config.distance_rainure_chant}mm",
            ])
        
        if self.config.decalage_arriere > 0:
            lines.append(f"  Décalage arrière: {self.config.decalage_arriere}mm")
        
        if self.config.hauteur_imposee:
            lines.append(f"  Hauteur imposée: {self.config.hauteur_imposee}mm ({self.config.position_verticale})")
        
        return "\n".join(lines)


# ============================================================================
# COMPARTIMENT
# ============================================================================

@dataclass
class Compartiment:
    """
    Représente un compartiment du meuble (espace entre séparations).
    
    Chaque compartiment peut recevoir:
    - Des étagères
    - Une ou plusieurs portes
    - Des tiroirs
    """
    index: int                      # Numéro du compartiment (0 = gauche)
    position_y: float               # Position Y du début du compartiment
    largeur: float                  # Largeur intérieure du compartiment
    hauteur: float                  # Hauteur intérieure
    profondeur: float               # Profondeur (depuis le fond)
    position_z_bas: float           # Position Z du bas du compartiment
    
    # Éléments du compartiment
    etageres: List[float] = field(default_factory=list)  # Hauteurs des étagères
    portes: List['Porte'] = field(default_factory=list)
    tiroirs: List['TiroirLegrabox'] = field(default_factory=list)
    
    def __repr__(self):
        return f"Compartiment({self.index}, y={self.position_y:.0f}, l={self.largeur:.0f})"


# ============================================================================
# SÉPARATION VERTICALE
# ============================================================================

@dataclass
class SeparationVerticale:
    """
    Panneau de séparation verticale à l'intérieur du meuble.
    
    Assemblage par vissage:
    - Vissé dans le panneau du dessous
    - Vissé dans la traverse haute (ou dessus)
    - Positionné contre le fond ou en retrait
    """
    nom: str
    hauteur: float                  # Hauteur de la séparation
    profondeur: float               # Profondeur (depuis l'arrière)
    epaisseur: float = 19
    position: Tuple[float, float, float] = (0, 0, 0)
    
    # Retrait par rapport au fond
    retrait_arriere: float = 0      # Distance depuis le fond
    
    # Retrait par rapport à l'avant
    retrait_avant: float = 0        # Distance depuis l'avant
    
    materiau: Optional[Materiau] = None
    couleur: Tuple[float, float, float] = (0.82, 0.71, 0.55)
    
    # Chants
    chants: Dict[str, float] = field(default_factory=dict)
    couleur_chant: Tuple[float, float, float] = (0.4, 0.4, 0.4)
    
    # Rainures pour crémaillères (2 par séparation: gauche et droite)
    rainures_cremaillere: List[Dict] = field(default_factory=list)
    
    def ajouter_chant(self, face: str, epaisseur: float = 1) -> 'SeparationVerticale':
        """Ajoute un chant sur une face ('avant', 'arriere', 'haut', 'bas')"""
        self.chants[face] = epaisseur
        return self
    
    def chant_avant(self, epaisseur: float = 1) -> 'SeparationVerticale':
        """Ajoute un chant sur la face avant uniquement"""
        self.chants['avant'] = epaisseur
        return self
    
    def ajouter_rainure_cremaillere(
        self,
        distance_avant: float,
        largeur: float = 16.0,
        profondeur: float = 7.0,
        face: str = "gauche"        # "gauche" ou "droite"
    ) -> 'SeparationVerticale':
        """Ajoute une rainure pour crémaillère sur une face"""
        self.rainures_cremaillere.append({
            'distance_avant': distance_avant,
            'largeur': largeur,
            'profondeur': profondeur,
            'face': face
        })
        return self
    
    def get_surface(self) -> float:
        return (self.profondeur * self.hauteur) / 1_000_000
    
    def get_lineaire_chant(self) -> Dict[str, float]:
        result = {}
        for face in self.chants.keys():
            if face in ['avant', 'arriere']:
                result[face] = self.hauteur / 1000
            else:  # haut, bas
                result[face] = self.profondeur / 1000
        return result
    
    def construire(self, gestionnaire: GestionnaireObjets, prefixe: str = "") -> List:
        objets = []
        nom_complet = f"{prefixe}_{self.nom}" if prefixe else self.nom
        couleur = self.materiau.couleur if self.materiau else self.couleur
        ep = self.materiau.epaisseur if self.materiau else self.epaisseur
        
        # Le panneau est orienté XZ (vertical, perpendiculaire à Y)
        shape = Part.makeBox(self.profondeur, ep, self.hauteur)
        shape.translate(App.Vector(*self.position))
        
        # Soustraire les rainures de crémaillères
        for rainure in self.rainures_cremaillere:
            rainure_shape = self._creer_shape_rainure_cremaillere(rainure, ep)
            if rainure_shape:
                try:
                    shape = shape.cut(rainure_shape)
                except:
                    pass
        
        obj = gestionnaire.get_ou_creer_objet(nom_complet)
        gestionnaire.mettre_a_jour_shape(obj, shape)
        gestionnaire.appliquer_couleur(obj, couleur)
        objets.append(obj)
        
        # Chants
        for face, ep_chant in self.chants.items():
            chant_obj = self._construire_chant(gestionnaire, nom_complet, face, ep_chant, ep)
            if chant_obj:
                objets.append(chant_obj)
        
        return objets
    
    def _creer_shape_rainure_cremaillere(self, rainure: Dict, ep_panneau: float) -> Part.Shape:
        """Crée le shape de la rainure de crémaillère pour une séparation"""
        distance_avant = rainure['distance_avant']
        largeur = rainure['largeur']
        profondeur = rainure['profondeur']
        face = rainure.get('face', 'gauche')
        
        try:
            # Position X: depuis l'avant de la séparation
            x_rainure = self.position[0] + self.profondeur - distance_avant - largeur
            
            # Position Y: selon la face (gauche ou droite de la séparation)
            if face == "gauche":
                y_rainure = self.position[1]
            else:  # droite
                y_rainure = self.position[1] + ep_panneau - profondeur
            
            z_rainure = self.position[2]
            
            rainure_shape = Part.makeBox(largeur, profondeur, self.hauteur)
            rainure_shape.translate(App.Vector(x_rainure, y_rainure, z_rainure))
            
            return rainure_shape
            
        except Exception as e:
            print(f"Erreur création rainure crémaillère pour {self.nom}: {e}")
            return None
    
    def _construire_chant(self, gestionnaire, nom_sep, face, ep_chant, ep_panneau):
        pos = list(self.position)
        nom_chant = f"{nom_sep}_chant_{face}"
        
        if face == 'avant':
            # Chant sur la face avant (X+)
            chant_dims = (ep_chant, ep_panneau, self.hauteur)
            pos[0] += self.profondeur
        elif face == 'arriere':
            # Chant sur la face arrière (X-)
            chant_dims = (ep_chant, ep_panneau, self.hauteur)
            pos[0] -= ep_chant
        elif face == 'haut':
            # Chant sur le dessus
            chant_dims = (self.profondeur, ep_panneau, ep_chant)
            pos[2] += self.hauteur
        elif face == 'bas':
            # Chant sur le dessous
            chant_dims = (self.profondeur, ep_panneau, ep_chant)
            pos[2] -= ep_chant
        else:
            return None
        
        shape = Part.makeBox(*chant_dims)
        shape.translate(App.Vector(*pos))
        
        obj = gestionnaire.get_ou_creer_objet(nom_chant)
        gestionnaire.mettre_a_jour_shape(obj, shape)
        gestionnaire.appliquer_couleur(obj, self.couleur_chant)
        
        return obj


# ============================================================================
# PORTE
# ============================================================================

@dataclass
class Porte:
    nom: str
    largeur: float
    hauteur: float
    epaisseur: float = 19
    type_ouverture: TypeOuverture = TypeOuverture.GAUCHE
    position: Tuple[float, float, float] = (0, 0, 0)
    couleur: Tuple[float, float, float] = (0.9, 0.9, 0.9)
    charnieres: List[Dict] = field(default_factory=list)
    chants: Dict[str, float] = field(default_factory=dict)
    couleur_chant: Tuple[float, float, float] = (0.4, 0.4, 0.4)
    
    def configurer_charnieres(self, regles: ReglesAssemblage):
        positions_z = ClipTopSpec.positions_charnieres(self.hauteur)
        centre_depuis_bord = ClipTopSpec.centre_cuvette_depuis_bord()
        
        self.charnieres = []
        for pos_z in positions_z:
            if self.type_ouverture == TypeOuverture.GAUCHE:
                pos_y = centre_depuis_bord
            else:
                pos_y = self.largeur - centre_depuis_bord
            
            self.charnieres.append({
                'pos_y': pos_y,
                'pos_z': pos_z,
                'diametre': ClipTopSpec.diametre_cuvette,
                'profondeur': ClipTopSpec.profondeur_cuvette
            })
        return self
    
    def chants_tous(self, epaisseur: float = 1) -> 'Porte':
        for face in ['haut', 'bas', 'gauche', 'droite']:
            self.chants[face] = epaisseur
        return self
    
    def construire(self, gestionnaire: GestionnaireObjets, prefixe: str = "") -> List:
        objets = []
        nom_complet = f"{prefixe}_{self.nom}" if prefixe else self.nom
        
        shape = Part.makeBox(self.epaisseur, self.largeur, self.hauteur)
        shape.translate(App.Vector(*self.position))
        
        obj = gestionnaire.get_ou_creer_objet(nom_complet)
        gestionnaire.mettre_a_jour_shape(obj, shape)
        gestionnaire.appliquer_couleur(obj, self.couleur, transparence=10)
        objets.append(obj)
        
        # Charnières (cuvettes usinées sur la face intérieure de la porte)
        for i, ch in enumerate(self.charnieres):
            cyl = Part.makeCylinder(ch['diametre'] / 2, ch['profondeur'])
            # La cuvette est usinée depuis la face intérieure (côté meuble)
            # Face intérieure = position X de la porte (self.position[0])
            pos_x = self.position[0]
            pos_y = self.position[1] + ch['pos_y']
            pos_z = self.position[2] + ch['pos_z']
            cyl.translate(App.Vector(pos_x, pos_y, pos_z))
            cyl.rotate(App.Vector(pos_x, pos_y, pos_z), App.Vector(0, 1, 0), 90)
            
            obj_ch = gestionnaire.get_ou_creer_objet(f"{nom_complet}_charniere_{i+1}")
            gestionnaire.mettre_a_jour_shape(obj_ch, cyl)
            gestionnaire.appliquer_couleur(obj_ch, (0.8, 0.6, 0.2))
            objets.append(obj_ch)
        
        # Chants
        for face, ep_chant in self.chants.items():
            pos = list(self.position)
            nom_chant = f"{nom_complet}_chant_{face}"
            
            if face == 'haut':
                chant_dims = (self.epaisseur, self.largeur, ep_chant)
                pos[2] += self.hauteur
            elif face == 'bas':
                chant_dims = (self.epaisseur, self.largeur, ep_chant)
                pos[2] -= ep_chant
            elif face == 'gauche':
                chant_dims = (self.epaisseur, ep_chant, self.hauteur)
                pos[1] -= ep_chant
            elif face == 'droite':
                chant_dims = (self.epaisseur, ep_chant, self.hauteur)
                pos[1] += self.largeur
            else:
                continue
            
            shape_ch = Part.makeBox(*chant_dims)
            shape_ch.translate(App.Vector(*pos))
            
            obj_ch = gestionnaire.get_ou_creer_objet(nom_chant)
            gestionnaire.mettre_a_jour_shape(obj_ch, shape_ch)
            gestionnaire.appliquer_couleur(obj_ch, self.couleur_chant)
            objets.append(obj_ch)
        
        return objets


# ============================================================================
# TIROIR LEGRABOX
# ============================================================================

@dataclass
class TiroirLegrabox:
    nom: str
    largeur_facade: float               # Largeur de la façade (calculée comme les portes)
    hauteur_facade: float               # Hauteur de la façade
    largeur_interieure: float           # Largeur intérieure du compartiment (entre panneaux)
    profondeur_corps: float             # Profondeur du corps du meuble
    hauteur_cote: LegraboxHauteur = LegraboxHauteur.M
    ep_facade: float = 19
    position: Tuple[float, float, float] = (0, 0, 0)
    couleur_facade: Tuple[float, float, float] = (0.9, 0.9, 0.9)
    couleur_legrabox: Tuple[float, float, float] = (0.7, 0.7, 0.72)
    couleur_fond: Tuple[float, float, float] = (0.85, 0.85, 0.85)
    
    @property
    def largeur_tiroir(self) -> float:
        return self.largeur_interieure - LegraboxSpec.jeu_total_lateral - 2 * LegraboxSpec.epaisseur_paroi
    
    @property
    def longueur_coulisse(self) -> int:
        return LegraboxSpec.calculer_profondeur_tiroir(self.profondeur_corps, self.ep_facade)
    
    @property
    def hauteur_paroi(self) -> float:
        return LegraboxSpec.get_hauteur_cote(self.hauteur_cote)
    
    def construire(self, gestionnaire: GestionnaireObjets, prefixe: str = "") -> List:
        objets = []
        nom_base = f"{prefixe}_{self.nom}" if prefixe else self.nom
        
        # Façade (positionnée comme une porte)
        facade = Part.makeBox(self.ep_facade, self.largeur_facade, self.hauteur_facade)
        facade.translate(App.Vector(*self.position))
        obj_facade = gestionnaire.get_ou_creer_objet(f"{nom_base}_facade")
        gestionnaire.mettre_a_jour_shape(obj_facade, facade)
        gestionnaire.appliquer_couleur(obj_facade, self.couleur_facade)
        objets.append(obj_facade)
        
        # Caisson du tiroir — centré dans le compartiment
        # pos_y du caisson = centre du compartiment - demi-largeur tiroir totale
        centre_facade_y = self.position[1] + self.largeur_facade / 2
        largeur_totale_tiroir = self.largeur_interieure - LegraboxSpec.jeu_total_lateral
        pos_tiroir_y = centre_facade_y - largeur_totale_tiroir / 2
        
        longueur = self.longueur_coulisse - 20
        hauteur_paroi = self.hauteur_paroi
        # Le caisson part de la façade vers l'arrière du meuble (X décroissant)
        pos_tiroir_x = self.position[0] - 3 - longueur
        pos_tiroir_z = self.position[2] + 10
        
        # Paroi gauche
        paroi_g = Part.makeBox(longueur, LegraboxSpec.epaisseur_paroi, hauteur_paroi)
        paroi_g.translate(App.Vector(pos_tiroir_x, pos_tiroir_y, pos_tiroir_z))
        obj_pg = gestionnaire.get_ou_creer_objet(f"{nom_base}_paroi_g")
        gestionnaire.mettre_a_jour_shape(obj_pg, paroi_g)
        gestionnaire.appliquer_couleur(obj_pg, self.couleur_legrabox)
        objets.append(obj_pg)
        
        # Paroi droite
        pos_paroi_d_y = pos_tiroir_y + self.largeur_tiroir + LegraboxSpec.epaisseur_paroi
        paroi_d = Part.makeBox(longueur, LegraboxSpec.epaisseur_paroi, hauteur_paroi)
        paroi_d.translate(App.Vector(pos_tiroir_x, pos_paroi_d_y, pos_tiroir_z))
        obj_pd = gestionnaire.get_ou_creer_objet(f"{nom_base}_paroi_d")
        gestionnaire.mettre_a_jour_shape(obj_pd, paroi_d)
        gestionnaire.appliquer_couleur(obj_pd, self.couleur_legrabox)
        objets.append(obj_pd)
        
        # Fond
        fond = Part.makeBox(longueur - 10, self.largeur_tiroir, LegraboxSpec.epaisseur_fond_standard)
        fond.translate(App.Vector(pos_tiroir_x + 5, pos_tiroir_y + LegraboxSpec.epaisseur_paroi, pos_tiroir_z))
        obj_fond = gestionnaire.get_ou_creer_objet(f"{nom_base}_fond")
        gestionnaire.mettre_a_jour_shape(obj_fond, fond)
        gestionnaire.appliquer_couleur(obj_fond, self.couleur_fond)
        objets.append(obj_fond)
        
        # Arrière
        arriere = Part.makeBox(LegraboxSpec.epaisseur_paroi, self.largeur_tiroir, hauteur_paroi)
        arriere.translate(App.Vector(
            pos_tiroir_x + longueur - LegraboxSpec.epaisseur_paroi - 5,
            pos_tiroir_y + LegraboxSpec.epaisseur_paroi,
            pos_tiroir_z
        ))
        obj_arr = gestionnaire.get_ou_creer_objet(f"{nom_base}_arriere")
        gestionnaire.mettre_a_jour_shape(obj_arr, arriere)
        gestionnaire.appliquer_couleur(obj_arr, self.couleur_legrabox)
        objets.append(obj_arr)
        
        return objets


# ============================================================================
# MEUBLE PRINCIPAL
# ============================================================================

class Meuble:
    """
    Classe principale avec gestion avancée du fond et des séparations.
    
    Les séparations verticales divisent le meuble en compartiments.
    Chaque compartiment peut recevoir des étagères, portes ou tiroirs.
    """
    
    def __init__(
        self,
        nom: str,
        largeur: float,
        hauteur: float,
        profondeur: float,
        epaisseur: float = 19,
        hauteur_plinthe: float = 100,
        epaisseur_facade: float = 19,
        materiau: Optional[Materiau] = None,
        regles: Optional[ReglesAssemblage] = None
    ):
        self.nom = nom
        self.largeur = largeur
        self.hauteur = hauteur
        self.profondeur = profondeur
        self.epaisseur = epaisseur
        self.hauteur_plinthe = hauteur_plinthe
        self.epaisseur_facade = epaisseur_facade
        self.materiau = materiau or Materiau.melamine_chene()
        self.regles = regles or ReglesAssemblage.standard()
        
        self._calculer_dimensions()
        
        # Structure
        self.panneaux: List[Panneau] = []
        self.fond: Optional[PanneauFond] = None
        
        # Séparations et compartiments
        self.separations: List[SeparationVerticale] = []
        self.compartiments: List[Compartiment] = []
        
        # Façades
        self.portes: List[Porte] = []
        self.tiroirs: List[TiroirLegrabox] = []
        
        # Pieds et plinthes
        self.pieds: List[Dict] = []
        self.plinthes: List[Dict] = []
        
        self.doc = None
        self.gestionnaire = None
    
    def _calculer_dimensions(self):
        ep = self.materiau.epaisseur
        self.profondeur_corps = self.profondeur - self.epaisseur_facade
        
        if self.regles.type_assemblage == TypeAssemblage.DESSUS_ENTRE:
            self.hauteur_corps = self.hauteur - self.hauteur_plinthe
            self.hauteur_cote = self.hauteur_corps
            self.largeur_dessus = self.largeur - 2 * ep
            self.position_dessus_y = ep
        else:
            self.hauteur_corps = self.hauteur - self.hauteur_plinthe
            self.hauteur_cote = self.hauteur_corps - 2 * ep
            self.largeur_dessus = self.largeur
            self.position_dessus_y = 0
        
        self.largeur_interieure = self.largeur - 2 * ep
    
    # -------------------------------------------------------------------------
    # STRUCTURE
    # -------------------------------------------------------------------------
    
    def ajouter_cotes(self, chant_avant: bool = True, ep_chant: float = 1) -> 'Meuble':
        ep = self.materiau.epaisseur
        config_fond = self.regles.config_fond
        
        if self.regles.type_assemblage == TypeAssemblage.DESSUS_ENTRE:
            hauteur_cote = self.hauteur_corps
            pos_z = self.hauteur_plinthe
        else:
            hauteur_cote = self.hauteur_corps - 2 * ep
            pos_z = self.hauteur_plinthe + ep
        
        # Gauche
        pan_g = Panneau(
            nom="Cote_Gauche",
            longueur=self.profondeur_corps,
            largeur=hauteur_cote,
            epaisseur=ep,
            position=(0, 0, pos_z),
            orientation="XZ",
            materiau=self.materiau
        )
        if chant_avant:
            pan_g.ajouter_chant('avant', ep_chant)
        
        # Ajouter rainure si assemblage par rainure
        if config_fond.type_assemblage == TypeAssemblageFond.RAINURE:
            pan_g.ajouter_rainure(
                position=config_fond.distance_rainure_chant + config_fond.decalage_arriere,
                largeur=config_fond.epaisseur + config_fond.jeu_fond_rainure,
                profondeur=config_fond.profondeur_rainure,
                direction="Z"
            )
        
        self.panneaux.append(pan_g)
        
        # Droit
        pan_d = Panneau(
            nom="Cote_Droit",
            longueur=self.profondeur_corps,
            largeur=hauteur_cote,
            epaisseur=ep,
            position=(0, self.largeur - ep, pos_z),
            orientation="XZ",
            materiau=self.materiau
        )
        if chant_avant:
            pan_d.ajouter_chant('avant', ep_chant)
        
        if config_fond.type_assemblage == TypeAssemblageFond.RAINURE:
            pan_d.ajouter_rainure(
                position=config_fond.distance_rainure_chant + config_fond.decalage_arriere,
                largeur=config_fond.epaisseur + config_fond.jeu_fond_rainure,
                profondeur=config_fond.profondeur_rainure,
                direction="Z"
            )
        
        self.panneaux.append(pan_d)
        
        return self
    
    def ajouter_dessous(self, retrait: float = 50, chant_avant: bool = True, ep_chant: float = 1) -> 'Meuble':
        ep = self.materiau.epaisseur
        config_fond = self.regles.config_fond
        
        pan = Panneau(
            nom="Dessous",
            longueur=self.profondeur_corps - retrait,
            largeur=self.largeur_dessus,
            epaisseur=ep,
            position=(retrait, self.position_dessus_y, self.hauteur_plinthe),
            orientation="XY",
            materiau=self.materiau
        )
        if chant_avant:
            pan.ajouter_chant('avant', ep_chant)
        
        # Rainure pour fond
        if config_fond.type_assemblage == TypeAssemblageFond.RAINURE:
            # Position de la rainure depuis le bord arrière du panneau (qui est à X=retrait)
            # La rainure doit être à distance_rainure_chant depuis l'arrière du MEUBLE (X=0)
            # Donc depuis le bord du panneau: distance_rainure_chant + decalage - retrait
            # Mais le panneau commence à retrait, donc la position relative est:
            pos_rainure_absolue = config_fond.distance_rainure_chant + config_fond.decalage_arriere
            pos_rainure_relative = pos_rainure_absolue - retrait
            
            if pos_rainure_relative >= 0 and pos_rainure_relative < (self.profondeur_corps - retrait):
                pan.ajouter_rainure(
                    position=pos_rainure_relative,
                    largeur=config_fond.epaisseur + config_fond.jeu_fond_rainure,
                    profondeur=config_fond.profondeur_rainure,
                    direction="Y"
                )
        
        self.panneaux.append(pan)
        return self
    
    def ajouter_dessus(
        self,
        type: str = "traverses",
        largeur_traverse: float = 100,
        chant: bool = True,
        ep_chant: float = 1
    ) -> 'Meuble':
        ep = self.materiau.epaisseur
        config_fond = self.regles.config_fond
        
        if self.regles.type_assemblage == TypeAssemblage.DESSUS_ENTRE:
            z = self.hauteur_plinthe + self.hauteur_corps - ep
        else:
            z = self.hauteur - ep
        
        if type == "plein":
            pan = Panneau(
                nom="Dessus",
                longueur=self.profondeur_corps,
                largeur=self.largeur_dessus,
                epaisseur=ep,
                position=(0, self.position_dessus_y, z),
                orientation="XY",
                materiau=self.materiau
            )
            if chant:
                pan.ajouter_chant('avant', ep_chant)
            
            # Rainure pour fond
            if config_fond.type_assemblage == TypeAssemblageFond.RAINURE:
                pan.ajouter_rainure(
                    position=config_fond.distance_rainure_chant + config_fond.decalage_arriere,
                    largeur=config_fond.epaisseur + config_fond.jeu_fond_rainure,
                    profondeur=config_fond.profondeur_rainure,
                    direction="Y"
                )
            
            self.panneaux.append(pan)
        else:
            # Traverse avant
            pan_av = Panneau(
                nom="Traverse_Avant",
                longueur=largeur_traverse,
                largeur=self.largeur_dessus,
                epaisseur=ep,
                position=(self.profondeur_corps - largeur_traverse, self.position_dessus_y, z),
                orientation="XY",
                materiau=self.materiau
            )
            if chant:
                pan_av.ajouter_chant('avant', ep_chant)
            self.panneaux.append(pan_av)
            
            # Traverse arrière
            pan_ar = Panneau(
                nom="Traverse_Arriere",
                longueur=largeur_traverse,
                largeur=self.largeur_dessus,
                epaisseur=ep,
                position=(0, self.position_dessus_y, z),
                orientation="XY",
                materiau=self.materiau
            )
            if chant:
                pan_ar.ajouter_chant('arriere', ep_chant)
            
            # Rainure dans traverse arrière
            if config_fond.type_assemblage == TypeAssemblageFond.RAINURE:
                pan_ar.ajouter_rainure(
                    position=config_fond.distance_rainure_chant + config_fond.decalage_arriere,
                    largeur=config_fond.epaisseur + config_fond.jeu_fond_rainure,
                    profondeur=config_fond.profondeur_rainure,
                    direction="Y"
                )
            
            self.panneaux.append(pan_ar)
        
        return self
    
    def ajouter_fond(
        self,
        config: ConfigFond = None,
        epaisseur: float = None,
        materiau: Materiau = None
    ) -> 'Meuble':
        """
        Ajoute le panneau de fond avec la configuration spécifiée.
        
        Args:
            config: Configuration du fond (si None, utilise celle des règles)
            epaisseur: Épaisseur du fond (surcharge la config)
            materiau: Matériau du fond
        """
        ep_cote = self.materiau.epaisseur
        cfg = config or self.regles.config_fond
        ep_fond = epaisseur or cfg.epaisseur
        
        # Calculer les dimensions selon le type d'assemblage
        if cfg.type_assemblage == TypeAssemblageFond.VISSAGE:
            # Fond vissé entre les panneaux
            largeur_fond = self.largeur_interieure
            
            if cfg.hauteur_imposee:
                hauteur_fond = cfg.hauteur_imposee
            else:
                hauteur_fond = self.hauteur_corps - 2 * ep_cote
            
            pos_y = ep_cote
            
            # Position Z selon placement vertical
            if cfg.hauteur_imposee:
                if cfg.position_verticale == "haut":
                    pos_z = self.hauteur_plinthe + self.hauteur_corps - ep_cote - hauteur_fond
                elif cfg.position_verticale == "centre":
                    espace = self.hauteur_corps - 2 * ep_cote - hauteur_fond
                    pos_z = self.hauteur_plinthe + ep_cote + espace / 2
                else:  # bas
                    pos_z = self.hauteur_plinthe + ep_cote
            else:
                pos_z = self.hauteur_plinthe + ep_cote
            
            pos_x = cfg.decalage_arriere
            
        elif cfg.type_assemblage == TypeAssemblageFond.RAINURE:
            # Fond embrèvé dans les rainures
            # Largeur = largeur intérieure + 2 * profondeur rainure - jeu
            largeur_fond = self.largeur_interieure + 2 * cfg.profondeur_rainure - cfg.jeu_fond_rainure
            
            if cfg.hauteur_imposee:
                # Hauteur imposée + embrèvement haut et bas
                hauteur_fond = cfg.hauteur_imposee + 2 * cfg.profondeur_rainure - cfg.jeu_fond_rainure
            else:
                # Hauteur totale avec embrèvements
                hauteur_fond = self.hauteur_corps - 2 * ep_cote + 2 * cfg.profondeur_rainure - cfg.jeu_fond_rainure
            
            # Position Y: centré, le fond dépasse dans les rainures
            pos_y = ep_cote - cfg.profondeur_rainure + cfg.jeu_fond_rainure / 2
            
            # Position Z
            if cfg.hauteur_imposee:
                if cfg.position_verticale == "haut":
                    pos_z = self.hauteur_plinthe + self.hauteur_corps - ep_cote - cfg.hauteur_imposee - cfg.profondeur_rainure + cfg.jeu_fond_rainure / 2
                elif cfg.position_verticale == "centre":
                    espace = self.hauteur_corps - 2 * ep_cote - cfg.hauteur_imposee
                    pos_z = self.hauteur_plinthe + ep_cote + espace / 2 - cfg.profondeur_rainure + cfg.jeu_fond_rainure / 2
                else:  # bas
                    pos_z = self.hauteur_plinthe + ep_cote - cfg.profondeur_rainure + cfg.jeu_fond_rainure / 2
            else:
                pos_z = self.hauteur_plinthe + ep_cote - cfg.profondeur_rainure + cfg.jeu_fond_rainure / 2
            
            pos_x = cfg.distance_rainure_chant + cfg.decalage_arriere
            
        else:  # APPLIQUE
            # Fond appliqué sur l'arrière
            largeur_fond = self.largeur
            
            if cfg.hauteur_imposee:
                hauteur_fond = cfg.hauteur_imposee
            else:
                hauteur_fond = self.hauteur_corps
            
            pos_y = 0
            
            if cfg.hauteur_imposee:
                if cfg.position_verticale == "haut":
                    pos_z = self.hauteur_plinthe + self.hauteur_corps - hauteur_fond
                elif cfg.position_verticale == "centre":
                    espace = self.hauteur_corps - hauteur_fond
                    pos_z = self.hauteur_plinthe + espace / 2
                else:
                    pos_z = self.hauteur_plinthe
            else:
                pos_z = self.hauteur_plinthe
            
            pos_x = cfg.decalage_arriere
        
        # Créer le panneau de fond
        self.fond = PanneauFond(
            nom="Fond",
            largeur=largeur_fond,
            hauteur=hauteur_fond,
            epaisseur=ep_fond,
            position=(pos_x, pos_y, pos_z),
            config=cfg,
            couleur=materiau.couleur if materiau else cfg.couleur
        )
        
        return self
    
    def ajouter_separations(
        self,
        config: ConfigSeparations = None,
        chant_avant: bool = True,
        ep_chant: float = 1
    ) -> 'Meuble':
        """
        Ajoute les séparations verticales et crée les compartiments.
        
        Les séparations divisent l'espace intérieur en compartiments.
        Assemblage par vissage dans le dessous et la traverse haute.
        
        Args:
            config: Configuration des séparations (si None, utilise celle des règles)
            chant_avant: Ajouter un chant sur la face avant
            ep_chant: Épaisseur du chant
        """
        cfg = config or self.regles.config_separations
        ep = self.materiau.epaisseur
        
        # Pas de séparation si 1 seul compartiment
        if cfg.nombre_compartiments <= 1 and not cfg.largeurs_compartiments:
            self._creer_compartiment_unique()
            return self
        
        # Calculer les dimensions des séparations
        # Hauteur: entre dessous et traverse haute
        hauteur_sep = self.hauteur_corps - 2 * ep
        
        # Profondeur: du fond (+ retrait) à l'avant (- retrait)
        config_fond = self.regles.config_fond
        
        # Position X de départ (depuis le fond)
        if config_fond.type_assemblage == TypeAssemblageFond.RAINURE:
            pos_x_depart = config_fond.distance_rainure_chant + config_fond.decalage_arriere + config_fond.epaisseur
        else:
            pos_x_depart = config_fond.decalage_arriere + config_fond.epaisseur
        
        pos_x_depart += cfg.retrait_arriere
        
        # Position X de fin (avant du corps - retrait)
        pos_x_fin = self.profondeur_corps - cfg.retrait_avant
        
        profondeur_sep = pos_x_fin - pos_x_depart
        
        # Position Z (sur le dessous)
        pos_z = self.hauteur_plinthe + ep
        
        # Calculer les largeurs des compartiments
        if cfg.largeurs_compartiments:
            # Mode largeurs personnalisées
            largeurs = cfg.largeurs_compartiments
            nb_sep = len(largeurs) - 1
        else:
            # Mode compartiments égaux
            nb_sep = cfg.nombre_compartiments - 1
            largeur_totale = self.largeur_interieure - nb_sep * cfg.epaisseur
            largeur_compartiment = largeur_totale / cfg.nombre_compartiments
            largeurs = [largeur_compartiment] * cfg.nombre_compartiments
        
        # Créer les séparations et les compartiments
        self.separations = []
        self.compartiments = []
        
        pos_y = ep  # Position Y courante (commence après le côté gauche)
        
        for i in range(cfg.nombre_compartiments):
            largeur_comp = largeurs[i]
            
            # Créer le compartiment
            compartiment = Compartiment(
                index=i,
                position_y=pos_y,
                largeur=largeur_comp,
                hauteur=hauteur_sep,
                profondeur=profondeur_sep,
                position_z_bas=pos_z
            )
            self.compartiments.append(compartiment)
            
            pos_y += largeur_comp
            
            # Créer la séparation (sauf après le dernier compartiment)
            if i < cfg.nombre_compartiments - 1:
                separation = SeparationVerticale(
                    nom=f"Separation_{i+1}",
                    hauteur=hauteur_sep,
                    profondeur=profondeur_sep,
                    epaisseur=cfg.epaisseur,
                    position=(pos_x_depart, pos_y, pos_z),
                    retrait_arriere=cfg.retrait_arriere,
                    retrait_avant=cfg.retrait_avant,
                    materiau=self.materiau
                )
                
                if chant_avant:
                    separation.chant_avant(ep_chant)
                
                self.separations.append(separation)
                pos_y += cfg.epaisseur
        
        return self
    
    def _creer_compartiment_unique(self):
        """Crée un compartiment unique (pas de séparation)"""
        ep = self.materiau.epaisseur
        config_fond = self.regles.config_fond
        
        # Calculer la profondeur disponible
        if config_fond.type_assemblage == TypeAssemblageFond.RAINURE:
            pos_x_depart = config_fond.distance_rainure_chant + config_fond.decalage_arriere + config_fond.epaisseur
        else:
            pos_x_depart = config_fond.decalage_arriere + config_fond.epaisseur
        
        profondeur_comp = self.profondeur_corps - pos_x_depart
        
        compartiment = Compartiment(
            index=0,
            position_y=ep,
            largeur=self.largeur_interieure,
            hauteur=self.hauteur_corps - 2 * ep,
            profondeur=profondeur_comp,
            position_z_bas=self.hauteur_plinthe + ep
        )
        self.compartiments = [compartiment]
    
    def ajouter_cremailleres(self, config: ConfigCremaillere = None) -> 'Meuble':
        """
        Ajoute les rainures pour crémaillères sur les côtés et séparations.
        
        Les crémaillères sont des profilés aluminium encastrés verticalement.
        Elles permettent de positionner les étagères à différentes hauteurs.
        
        Args:
            config: Configuration des crémaillères (si None, utilise celle des règles)
        """
        cfg = config or self.regles.config_cremaillere
        
        # Ajouter les rainures sur les côtés
        for panneau in self.panneaux:
            if "Cote_Gauche" in panneau.nom or "Cote_Droit" in panneau.nom:
                if cfg.double:
                    # Crémaillère avant
                    panneau.ajouter_rainure_cremaillere(
                        distance_avant=cfg.distance_avant,
                        largeur=cfg.largeur,
                        profondeur=cfg.profondeur
                    )
                    # Crémaillère arrière
                    panneau.ajouter_rainure_cremaillere(
                        distance_avant=panneau.longueur - cfg.distance_arriere - cfg.largeur,
                        largeur=cfg.largeur,
                        profondeur=cfg.profondeur
                    )
                else:
                    # Une seule crémaillère
                    if cfg.position_unique == "avant":
                        distance = cfg.distance_avant
                    else:
                        distance = panneau.longueur - cfg.distance_arriere - cfg.largeur
                    
                    panneau.ajouter_rainure_cremaillere(
                        distance_avant=distance,
                        largeur=cfg.largeur,
                        profondeur=cfg.profondeur
                    )
        
        # Ajouter les rainures sur les séparations (des deux côtés)
        for separation in self.separations:
            if cfg.double:
                # Crémaillère avant - face gauche
                separation.ajouter_rainure_cremaillere(
                    distance_avant=cfg.distance_avant,
                    largeur=cfg.largeur,
                    profondeur=cfg.profondeur,
                    face="gauche"
                )
                # Crémaillère arrière - face gauche
                separation.ajouter_rainure_cremaillere(
                    distance_avant=separation.profondeur - cfg.distance_arriere - cfg.largeur,
                    largeur=cfg.largeur,
                    profondeur=cfg.profondeur,
                    face="gauche"
                )
                # Crémaillère avant - face droite
                separation.ajouter_rainure_cremaillere(
                    distance_avant=cfg.distance_avant,
                    largeur=cfg.largeur,
                    profondeur=cfg.profondeur,
                    face="droite"
                )
                # Crémaillère arrière - face droite
                separation.ajouter_rainure_cremaillere(
                    distance_avant=separation.profondeur - cfg.distance_arriere - cfg.largeur,
                    largeur=cfg.largeur,
                    profondeur=cfg.profondeur,
                    face="droite"
                )
            else:
                if cfg.position_unique == "avant":
                    distance = cfg.distance_avant
                else:
                    distance = separation.profondeur - cfg.distance_arriere - cfg.largeur
                
                separation.ajouter_rainure_cremaillere(
                    distance_avant=distance,
                    largeur=cfg.largeur,
                    profondeur=cfg.profondeur,
                    face="gauche"
                )
                separation.ajouter_rainure_cremaillere(
                    distance_avant=distance,
                    largeur=cfg.largeur,
                    profondeur=cfg.profondeur,
                    face="droite"
                )
        
        return self
    
    def ajouter_etagere_compartiment(
        self,
        compartiment_index: int,
        hauteur_depuis_bas: float,
        config: ConfigEtagere = None
    ) -> 'Meuble':
        """
        Ajoute une étagère sur crémaillères dans un compartiment.
        
        L'étagère utilise les paramètres de ConfigEtagere pour:
        - Le jeu latéral avec les côtés/séparations
        - Le jeu arrière avec le fond
        - Le retrait avant
        - Les chants
        
        Args:
            compartiment_index: Index du compartiment (0 = gauche)
            hauteur_depuis_bas: Hauteur depuis le bas du compartiment
            config: Configuration de l'étagère (si None, utilise celle des règles)
        """
        if compartiment_index >= len(self.compartiments):
            print(f"Erreur: Compartiment {compartiment_index} n'existe pas")
            return self
        
        cfg = config or self.regles.config_etagere
        comp = self.compartiments[compartiment_index]
        ep = self.materiau.epaisseur
        
        # Position X: depuis le fond + jeu arrière
        config_fond = self.regles.config_fond
        if config_fond.type_assemblage == TypeAssemblageFond.RAINURE:
            pos_x = config_fond.distance_rainure_chant + config_fond.decalage_arriere + config_fond.epaisseur
        else:
            pos_x = config_fond.decalage_arriere + config_fond.epaisseur
        
        pos_x += cfg.jeu_arriere
        
        # Calculer la profondeur de l'étagère
        profondeur_etagere = comp.profondeur - cfg.jeu_arriere - cfg.retrait_avant
        
        # Largeur avec jeux latéraux
        largeur_etagere = comp.largeur - 2 * cfg.jeu_lateral
        
        pan = Panneau(
            nom=f"Etagere_C{compartiment_index}_{int(hauteur_depuis_bas)}",
            longueur=profondeur_etagere,
            largeur=largeur_etagere,
            epaisseur=ep,
            position=(pos_x, comp.position_y + cfg.jeu_lateral, comp.position_z_bas + hauteur_depuis_bas),
            orientation="XY",
            materiau=self.materiau
        )
        
        # Chants
        if cfg.chant_avant:
            pan.ajouter_chant('avant', cfg.epaisseur_chant)
        if cfg.chant_arriere:
            pan.ajouter_chant('arriere', cfg.epaisseur_chant)
        if cfg.chant_lateraux:
            pan.ajouter_chant('gauche', cfg.epaisseur_chant)
            pan.ajouter_chant('droite', cfg.epaisseur_chant)
        
        self.panneaux.append(pan)
        comp.etageres.append(hauteur_depuis_bas)
        
        return self
    
    def ajouter_etageres_compartiment(
        self,
        compartiment_index: int,
        hauteurs: List[float],
        config: ConfigEtagere = None
    ) -> 'Meuble':
        """
        Ajoute plusieurs étagères dans un compartiment.
        
        Args:
            compartiment_index: Index du compartiment
            hauteurs: Liste des hauteurs depuis le bas
            config: Configuration des étagères
        """
        for hauteur in hauteurs:
            self.ajouter_etagere_compartiment(compartiment_index, hauteur, config)
        return self
    
    def ajouter_etagere(
        self,
        hauteur_depuis_bas: float,
        retrait_avant: float = None,
        chant_avant: bool = True,
        ep_chant: float = 1
    ) -> 'Meuble':
        """Ajoute une étagère simple (sans compartiments)"""
        ep = self.materiau.epaisseur
        cfg = self.regles.config_etagere
        retrait = retrait_avant if retrait_avant is not None else cfg.retrait_avant
        jeu_lat = cfg.jeu_lateral
        
        pan = Panneau(
            nom=f"Etagere_{int(hauteur_depuis_bas)}",
            longueur=self.profondeur_corps - retrait - cfg.jeu_arriere,
            largeur=self.largeur_interieure - 2 * jeu_lat,
            epaisseur=ep,
            position=(cfg.jeu_arriere, ep + jeu_lat, self.hauteur_plinthe + hauteur_depuis_bas),
            orientation="XY",
            materiau=self.materiau
        )
        if chant_avant:
            pan.ajouter_chant('avant', ep_chant)
        self.panneaux.append(pan)
        
        return self
    
    # -------------------------------------------------------------------------
    # PIEDS ET PLINTHES
    # -------------------------------------------------------------------------
    
    def ajouter_pieds_et_plinthes(self, config: ConfigPlinthe = None) -> 'Meuble':
        """
        Ajoute les pieds réglables et les plinthes au meuble.
        
        Les pieds sont répartis sous le meuble selon une grille.
        Les plinthes se fixent sur les pieds et couvrent 1, 2 ou 3 côtés.
        
        Assemblage aux angles : coupes droites.
        La plinthe avant passe devant les plinthes latérales.
        Les plinthes latérales sont alignées avec le côté du meuble.
        Seule la plinthe avant est en retrait.
        
        Args:
            config: Configuration des pieds/plinthes (si None, utilise celle des règles)
        """
        cfg = config or self.regles.config_plinthe
        
        # Synchroniser hauteur pied avec hauteur plinthe du meuble
        cfg.hauteur_pied = self.hauteur_plinthe
        cfg.hauteur = self.hauteur_plinthe
        
        self.pieds = []
        self.plinthes = []
        
        # --- Calculer la grille de pieds ---
        nb_prof = cfg.nombre_pieds_profondeur
        nb_larg = cfg.nombre_pieds_largeur
        
        # Positions en profondeur (X)
        if nb_prof == 1:
            positions_x = [self.profondeur_corps / 2]
        else:
            x_min = cfg.marge_pied
            x_max = self.profondeur_corps - cfg.marge_pied
            if nb_prof == 2:
                positions_x = [x_min, x_max]
            else:
                pas_x = (x_max - x_min) / (nb_prof - 1)
                positions_x = [x_min + i * pas_x for i in range(nb_prof)]
        
        # Positions en largeur (Y)
        if nb_larg == 1:
            positions_y = [self.largeur / 2]
        else:
            y_min = cfg.marge_pied
            y_max = self.largeur - cfg.marge_pied
            if nb_larg == 2:
                positions_y = [y_min, y_max]
            else:
                pas_y = (y_max - y_min) / (nb_larg - 1)
                positions_y = [y_min + i * pas_y for i in range(nb_larg)]
        
        # Créer les pieds
        rayon = cfg.diametre_pied / 2
        for ix, px in enumerate(positions_x):
            for iy, py in enumerate(positions_y):
                self.pieds.append({
                    'nom': f"Pied_{ix}_{iy}",
                    'position': (px, py, 0),
                    'rayon': rayon,
                    'hauteur': cfg.hauteur_pied,
                    'couleur': cfg.couleur_pied
                })
        
        # --- Créer les plinthes (coupes droites) ---
        retrait = cfg.retrait
        retrait_lat = cfg.retrait_lateral
        ep = cfg.epaisseur
        h = cfg.hauteur
        
        has_avant = "avant" in cfg.cotes
        has_gauche = "gauche" in cfg.cotes
        has_droite = "droite" in cfg.cotes
        
        for cote in cfg.cotes:
            if cote == "avant":
                # Plinthe avant : passe devant les plinthes latérales
                # Le retrait est mesuré depuis la face avant de la façade
                pos_x_avant = self.profondeur - retrait - ep
                
                # Position Y et longueur : s'aligne avec la plinthe latérale
                # si elle est présente, sinon avec le côté du meuble
                # Côté gauche (Y début)
                if has_gauche:
                    pos_y_debut = retrait_lat  # alignée avec la face extérieure de la plinthe gauche
                else:
                    pos_y_debut = 0  # alignée avec le côté du meuble
                
                # Côté droit (Y fin)
                if has_droite:
                    pos_y_fin = self.largeur - retrait_lat  # alignée avec la face extérieure de la plinthe droite
                else:
                    pos_y_fin = self.largeur  # alignée avec le côté du meuble
                
                longueur_avant = pos_y_fin - pos_y_debut
                
                self.plinthes.append({
                    'nom': "Plinthe_Avant",
                    'longueur': longueur_avant,
                    'hauteur': h,
                    'epaisseur': ep,
                    'position': (pos_x_avant, pos_y_debut, 0),
                    'orientation': 'YZ',
                    'couleur': cfg.couleur_plinthe
                })
            
            elif cote == "gauche":
                # Plinthe gauche : en retrait latéral depuis le côté du meuble
                # S'arrête contre la plinthe avant (si présente)
                if has_avant:
                    pos_x_avant = self.profondeur - retrait - ep
                    profondeur_plinthe = pos_x_avant  # du mur jusqu'à la plinthe avant
                else:
                    profondeur_plinthe = self.profondeur - retrait
                
                self.plinthes.append({
                    'nom': "Plinthe_Gauche",
                    'longueur': profondeur_plinthe,
                    'hauteur': h,
                    'epaisseur': ep,
                    'position': (0, retrait_lat, 0),
                    'orientation': 'XZ',
                    'couleur': cfg.couleur_plinthe
                })
            
            elif cote == "droite":
                # Plinthe droite : en retrait latéral depuis le côté du meuble
                if has_avant:
                    pos_x_avant = self.profondeur - retrait - ep
                    profondeur_plinthe = pos_x_avant
                else:
                    profondeur_plinthe = self.profondeur - retrait
                
                self.plinthes.append({
                    'nom': "Plinthe_Droite",
                    'longueur': profondeur_plinthe,
                    'hauteur': h,
                    'epaisseur': ep,
                    'position': (0, self.largeur - retrait_lat - ep, 0),
                    'orientation': 'XZ',
                    'couleur': cfg.couleur_plinthe
                })
        
        return self
    
    # -------------------------------------------------------------------------
    # FAÇADES
    # -------------------------------------------------------------------------
    
    def _calculer_dimensions_porte(self) -> Tuple[float, float, float, float]:
        """Calcule largeur, hauteur, pos_y, pos_z d'une porte selon le type de pose.
        
        EN_APPLIQUE : recouvrement de 16mm sur les panneaux latéraux
        SEMI_APPLIQUE : recouvrement de 8mm sur les panneaux latéraux
        ENCLOISONNEE : porte entre les panneaux, face alignée avec le corps,
                       jeu de 4mm entre la porte et le panneau
        """
        ep = self.materiau.epaisseur
        r = self.regles
        spec = ClipTopSpec
        
        if r.type_pose == TypePose.EN_APPLIQUE:
            rec = spec.recouvrement_applique  # 16mm
            largeur = self.largeur_interieure + 2 * rec - 2 * r.jeu_porte_lateral
            hauteur = self.hauteur_corps - r.jeu_porte_haut - r.jeu_porte_bas
            pos_y = ep - rec + r.jeu_porte_lateral
            pos_z = self.hauteur_plinthe + r.jeu_porte_bas
            
        elif r.type_pose == TypePose.SEMI_APPLIQUE:
            rec = spec.recouvrement_semi_applique  # 8mm
            largeur = self.largeur_interieure + 2 * rec - 2 * r.jeu_porte_lateral
            hauteur = self.hauteur_corps - r.jeu_porte_haut - r.jeu_porte_bas
            pos_y = ep - rec + r.jeu_porte_lateral
            pos_z = self.hauteur_plinthe + r.jeu_porte_bas
            
        else:  # ENCLOISONNEE
            jeu_encl = spec.jeu_encloisonnee  # 4mm
            largeur = self.largeur_interieure - 2 * jeu_encl
            hauteur = self.hauteur_corps - 2 * ep - r.jeu_porte_haut - r.jeu_porte_bas
            pos_y = ep + jeu_encl
            pos_z = self.hauteur_plinthe + ep + r.jeu_porte_bas
        
        return largeur, hauteur, pos_y, pos_z
    
    def ajouter_porte(
        self,
        position: Position = Position.CENTRE,
        type_ouverture: TypeOuverture = TypeOuverture.GAUCHE,
        chants: bool = True,
        ep_chant: float = 1
    ) -> 'Meuble':
        largeur, hauteur, pos_y, pos_z = self._calculer_dimensions_porte()
        
        # Position X selon le type de pose
        if self.regles.type_pose == TypePose.ENCLOISONNEE:
            # Face avant porte alignée avec face avant corps
            pos_x = self.profondeur_corps - self.epaisseur_facade
        else:
            # En applique / semi-applique : porte plaquée sur la face avant du corps
            pos_x = self.profondeur_corps
        
        porte = Porte(
            nom=f"Porte_{position.value}",
            largeur=largeur,
            hauteur=hauteur,
            epaisseur=self.epaisseur_facade,
            type_ouverture=type_ouverture,
            position=(pos_x, pos_y, pos_z)
        )
        porte.configurer_charnieres(self.regles)
        if chants:
            porte.chants_tous(ep_chant)
        
        self.portes.append(porte)
        return self
    
    def ajouter_portes_doubles(self, chants: bool = True, ep_chant: float = 1) -> 'Meuble':
        largeur_totale, hauteur, pos_y_base, pos_z = self._calculer_dimensions_porte()
        jeu_entre = self.regles.jeu_entre_portes
        larg_porte = (largeur_totale - jeu_entre) / 2
        
        # Position X selon le type de pose
        if self.regles.type_pose == TypePose.ENCLOISONNEE:
            pos_x = self.profondeur_corps - self.epaisseur_facade
        else:
            pos_x = self.profondeur_corps
        
        porte_g = Porte(
            nom="Porte_Gauche",
            largeur=larg_porte,
            hauteur=hauteur,
            epaisseur=self.epaisseur_facade,
            type_ouverture=TypeOuverture.GAUCHE,
            position=(pos_x, pos_y_base, pos_z)
        )
        porte_g.configurer_charnieres(self.regles)
        if chants:
            porte_g.chants_tous(ep_chant)
        self.portes.append(porte_g)
        
        porte_d = Porte(
            nom="Porte_Droite",
            largeur=larg_porte,
            hauteur=hauteur,
            epaisseur=self.epaisseur_facade,
            type_ouverture=TypeOuverture.DROITE,
            position=(pos_x, pos_y_base + larg_porte + jeu_entre, pos_z)
        )
        porte_d.configurer_charnieres(self.regles)
        if chants:
            porte_d.chants_tous(ep_chant)
        self.portes.append(porte_d)
        
        return self
    
    def _calculer_porte_compartiment(self, index_comp: int) -> Tuple[float, float, float, float]:
        """Calcule largeur, hauteur, pos_y, pos_z d'une porte pour un compartiment donné."""
        if not self.compartiments:
            self._creer_compartiment_unique()
        
        r = self.regles
        ep = self.materiau.epaisseur
        spec = ClipTopSpec
        nb_comp = len(self.compartiments)
        ep_sep = r.config_separations.epaisseur
        comp = self.compartiments[index_comp]
        
        is_premier = (index_comp == 0)
        is_dernier = (index_comp == nb_comp - 1)
        
        if r.type_pose == TypePose.EN_APPLIQUE:
            rec = spec.recouvrement_applique
            
            if is_premier:
                rec_g = rec
                jeu_g = r.jeu_porte_lateral
            else:
                rec_g = ep_sep / 2
                jeu_g = r.jeu_entre_portes / 2
            
            if is_dernier:
                rec_d = rec
                jeu_d = r.jeu_porte_lateral
            else:
                rec_d = ep_sep / 2
                jeu_d = r.jeu_entre_portes / 2
            
            largeur = comp.largeur + rec_g + rec_d - jeu_g - jeu_d
            pos_y = comp.position_y - rec_g + jeu_g
            hauteur = self.hauteur_corps - r.jeu_porte_haut - r.jeu_porte_bas
            pos_z = self.hauteur_plinthe + r.jeu_porte_bas
            
        elif r.type_pose == TypePose.SEMI_APPLIQUE:
            rec = spec.recouvrement_semi_applique
            
            if is_premier:
                rec_g = rec
                jeu_g = r.jeu_porte_lateral
            else:
                rec_g = ep_sep / 2
                jeu_g = r.jeu_entre_portes / 2
            
            if is_dernier:
                rec_d = rec
                jeu_d = r.jeu_porte_lateral
            else:
                rec_d = ep_sep / 2
                jeu_d = r.jeu_entre_portes / 2
            
            largeur = comp.largeur + rec_g + rec_d - jeu_g - jeu_d
            pos_y = comp.position_y - rec_g + jeu_g
            hauteur = self.hauteur_corps - r.jeu_porte_haut - r.jeu_porte_bas
            pos_z = self.hauteur_plinthe + r.jeu_porte_bas
            
        else:  # ENCLOISONNEE
            jeu_encl = spec.jeu_encloisonnee
            jeu_g = jeu_encl if is_premier else r.jeu_entre_portes / 2
            jeu_d = jeu_encl if is_dernier else r.jeu_entre_portes / 2
            
            largeur = comp.largeur - jeu_g - jeu_d
            pos_y = comp.position_y + jeu_g
            hauteur = self.hauteur_corps - 2 * ep - r.jeu_porte_haut - r.jeu_porte_bas
            pos_z = self.hauteur_plinthe + ep + r.jeu_porte_bas
        
        return largeur, hauteur, pos_y, pos_z
    
    def ajouter_portes_compartiments(
        self,
        portes: Optional[List[Tuple[int, TypeOuverture]]] = None,
        chants: bool = True,
        ep_chant: float = 1
    ) -> 'Meuble':
        """Ajoute une porte par compartiment.
        
        Args:
            portes: Liste de (n° compartiment, ouverture) pour chaque porte.
                    Si None, crée une porte par compartiment avec alternance G/D.
                    Exemples :
                        [(0, TypeOuverture.GAUCHE), (1, TypeOuverture.DROITE)]
                        [(0, TypeOuverture.GAUCHE), (2, TypeOuverture.DROITE)]  # pas de porte sur le 1
            chants: Appliquer des chants sur les portes
            ep_chant: Épaisseur des chants
        """
        if not self.compartiments:
            self._creer_compartiment_unique()
        
        # Par défaut : toutes les portes avec alternance G/D
        if portes is None:
            portes = [
                (i, TypeOuverture.GAUCHE if i % 2 == 0 else TypeOuverture.DROITE)
                for i in range(len(self.compartiments))
            ]
        
        # Position X selon le type de pose
        if self.regles.type_pose == TypePose.ENCLOISONNEE:
            pos_x = self.profondeur_corps - self.epaisseur_facade
        else:
            pos_x = self.profondeur_corps
        
        for index_comp, ouverture in portes:
            largeur_porte, hauteur, pos_y, pos_z = self._calculer_porte_compartiment(index_comp)
            
            porte = Porte(
                nom=f"Porte_{index_comp + 1}",
                largeur=largeur_porte,
                hauteur=hauteur,
                epaisseur=self.epaisseur_facade,
                type_ouverture=ouverture,
                position=(pos_x, pos_y, pos_z)
            )
            porte.configurer_charnieres(self.regles)
            if chants:
                porte.chants_tous(ep_chant)
            self.portes.append(porte)
        
        return self
    
    def ajouter_tiroirs_legrabox(
        self,
        nombre: int,
        index_compartiment: int = 0,
        hauteur_facade: float = None,
        hauteur_cote: LegraboxHauteur = None,
        chants: bool = True,
        ep_chant: float = 1
    ) -> 'Meuble':
        """Ajoute des tiroirs LEGRABOX identiques dans un compartiment.
        
        Args:
            nombre: Nombre de tiroirs
            index_compartiment: N° du compartiment (0 = gauche)
            hauteur_facade: Hauteur imposée (sinon calculée pour remplir)
            hauteur_cote: Hauteur des parois LEGRABOX
        """
        if not self.compartiments:
            self._creer_compartiment_unique()
        
        comp = self.compartiments[index_compartiment]
        r = self.regles
        h_cote = hauteur_cote or r.legrabox_hauteur
        
        largeur_facade_calc, hauteur_dispo, pos_y, pos_z_base = self._calculer_porte_compartiment(index_compartiment)
        
        # Position X : identique aux portes
        if r.type_pose == TypePose.ENCLOISONNEE:
            pos_x = self.profondeur_corps - self.epaisseur_facade
        else:
            pos_x = self.profondeur_corps
        
        jeu_entre = r.jeu_entre_tiroirs
        h_facade = hauteur_facade or (hauteur_dispo - (nombre - 1) * jeu_entre) / nombre
        
        pos_z = pos_z_base
        for i in range(nombre):
            tiroir = TiroirLegrabox(
                nom=f"Tiroir_{index_compartiment + 1}_{i+1}",
                largeur_facade=largeur_facade_calc,
                hauteur_facade=h_facade,
                largeur_interieure=comp.largeur,
                profondeur_corps=self.profondeur_corps,
                hauteur_cote=h_cote,
                ep_facade=self.epaisseur_facade,
                position=(pos_x, pos_y, pos_z)
            )
            self.tiroirs.append(tiroir)
            pos_z += h_facade + jeu_entre
        
        return self
    
    def ajouter_facades_compartiment(
        self,
        index_compartiment: int,
        facades: List[Tuple],
        chants: bool = True,
        ep_chant: float = 1
    ) -> 'Meuble':
        """Ajoute des façades (portes et/ou tiroirs) dans un compartiment.
        
        Les façades sont empilées de bas en haut avec un jeu de 4mm entre elles.
        Les façades sans hauteur imposée se répartissent l'espace disponible.
        
        Args:
            index_compartiment: N° du compartiment (0 = gauche)
            facades: Liste de tuples décrivant chaque façade, de bas en haut :
                ("porte", TypeOuverture)          — porte, hauteur auto
                ("porte", TypeOuverture, 400)     — porte, hauteur imposée
                ("tiroir", LegraboxHauteur)        — tiroir, hauteur façade auto
                ("tiroir", LegraboxHauteur, 200)   — tiroir, hauteur façade imposée
            
        Exemples :
            # 3 tiroirs égaux (hauteur auto)
            meuble.ajouter_facades_compartiment(0, [
                ("tiroir", LegraboxHauteur.K),
                ("tiroir", LegraboxHauteur.K),
                ("tiroir", LegraboxHauteur.K),
            ])
            
            # 1 tiroir K en bas + 1 porte en haut (hauteurs auto)
            meuble.ajouter_facades_compartiment(1, [
                ("tiroir", LegraboxHauteur.K),
                ("porte", TypeOuverture.GAUCHE),
            ])
        """
        if not self.compartiments:
            self._creer_compartiment_unique()
        
        comp = self.compartiments[index_compartiment]
        r = self.regles
        jeu_entre = r.jeu_entre_tiroirs  # 4mm entre façades
        
        largeur_facade_calc, hauteur_dispo, pos_y, pos_z_base = self._calculer_porte_compartiment(index_compartiment)
        
        # Position X
        if r.type_pose == TypePose.ENCLOISONNEE:
            pos_x = self.profondeur_corps - self.epaisseur_facade
        else:
            pos_x = self.profondeur_corps
        
        # Calculer les hauteurs
        nb_facades = len(facades)
        jeux_totaux = (nb_facades - 1) * jeu_entre
        
        # Séparer façades à hauteur imposée et façades auto
        hauteur_fixe = 0
        nb_auto = 0
        for f in facades:
            if len(f) > 2:
                hauteur_fixe += f[2]  # Hauteur imposée (3e élément)
            else:
                nb_auto += 1
        
        # Hauteur auto = espace restant réparti équitablement
        hauteur_restante = hauteur_dispo - jeux_totaux - hauteur_fixe
        h_auto = hauteur_restante / nb_auto if nb_auto > 0 else 0
        
        # Créer les façades de bas en haut
        pos_z = pos_z_base
        num_tiroir = 1
        num_porte = 1
        
        for f in facades:
            h_facade = f[2] if len(f) > 2 else h_auto
            
            if f[0] == "tiroir":
                h_cote = f[1] if len(f) > 1 else r.legrabox_hauteur
                
                tiroir = TiroirLegrabox(
                    nom=f"Tiroir_{index_compartiment + 1}_{num_tiroir}",
                    largeur_facade=largeur_facade_calc,
                    hauteur_facade=h_facade,
                    largeur_interieure=comp.largeur,
                    profondeur_corps=self.profondeur_corps,
                    hauteur_cote=h_cote,
                    ep_facade=self.epaisseur_facade,
                    position=(pos_x, pos_y, pos_z)
                )
                self.tiroirs.append(tiroir)
                num_tiroir += 1
                
            elif f[0] == "porte":
                ouverture = f[1] if len(f) > 1 else TypeOuverture.GAUCHE
                
                porte = Porte(
                    nom=f"Porte_{index_compartiment + 1}_{num_porte}",
                    largeur=largeur_facade_calc,
                    hauteur=h_facade,
                    epaisseur=self.epaisseur_facade,
                    type_ouverture=ouverture,
                    position=(pos_x, pos_y, pos_z)
                )
                porte.configurer_charnieres(self.regles)
                if chants:
                    porte.chants_tous(ep_chant)
                self.portes.append(porte)
                num_porte += 1
            
            pos_z += h_facade + jeu_entre
        
        return self
    
    # -------------------------------------------------------------------------
    # CONSTRUCTION
    # -------------------------------------------------------------------------
    
    def construire(self) -> 'Meuble':
        try:
            self.doc = App.getDocument(self.nom)
        except:
            self.doc = None
        
        if self.doc is None:
            self.doc = App.newDocument(self.nom)
        
        self.gestionnaire = GestionnaireObjets(self.doc)
        
        grp_racine = self.gestionnaire.get_ou_creer_groupe(self.nom)
        
        # Panneaux
        grp_panneaux = self.gestionnaire.get_ou_creer_groupe(f"{self.nom}_Panneaux", grp_racine)
        for panneau in self.panneaux:
            objets = panneau.construire(self.gestionnaire, self.nom)
            for obj in objets:
                if obj not in grp_panneaux.Group:
                    grp_panneaux.addObject(obj)
        
        # Fond
        if self.fond:
            obj_fond = self.fond.construire(self.gestionnaire, self.nom)
            if obj_fond not in grp_panneaux.Group:
                grp_panneaux.addObject(obj_fond)
        
        # Séparations verticales
        if self.separations:
            grp_separations = self.gestionnaire.get_ou_creer_groupe(f"{self.nom}_Separations", grp_racine)
            for separation in self.separations:
                objets = separation.construire(self.gestionnaire, self.nom)
                for obj in objets:
                    if obj not in grp_separations.Group:
                        grp_separations.addObject(obj)
        
        # Portes
        if self.portes:
            grp_portes = self.gestionnaire.get_ou_creer_groupe(f"{self.nom}_Portes", grp_racine)
            for porte in self.portes:
                objets = porte.construire(self.gestionnaire, self.nom)
                for obj in objets:
                    if obj not in grp_portes.Group:
                        grp_portes.addObject(obj)
        
        # Tiroirs
        if self.tiroirs:
            grp_tiroirs = self.gestionnaire.get_ou_creer_groupe(f"{self.nom}_Tiroirs", grp_racine)
            for tiroir in self.tiroirs:
                objets = tiroir.construire(self.gestionnaire, self.nom)
                for obj in objets:
                    if obj not in grp_tiroirs.Group:
                        grp_tiroirs.addObject(obj)
        
        # Pieds et plinthes
        if self.pieds or self.plinthes:
            grp_pieds = self.gestionnaire.get_ou_creer_groupe(f"{self.nom}_Pieds_Plinthes", grp_racine)
            
            # Pieds (cylindres)
            for pied in self.pieds:
                nom_pied = f"{self.nom}_{pied['nom']}"
                try:
                    shape = Part.makeCylinder(pied['rayon'], pied['hauteur'])
                    pos = pied['position']
                    shape.translate(App.Vector(pos[0], pos[1], pos[2]))
                    
                    obj = self.gestionnaire.get_ou_creer_objet(nom_pied)
                    self.gestionnaire.mettre_a_jour_shape(obj, shape)
                    self.gestionnaire.appliquer_couleur(obj, pied['couleur'])
                    if obj not in grp_pieds.Group:
                        grp_pieds.addObject(obj)
                except Exception as e:
                    print(f"Erreur construction pied {pied['nom']}: {e}")
            
            # Plinthes (panneaux — coupes droites)
            for plinthe in self.plinthes:
                nom_plinthe = f"{self.nom}_{plinthe['nom']}"
                try:
                    orientation = plinthe['orientation']
                    L = plinthe['longueur']
                    H = plinthe['hauteur']
                    ep = plinthe['epaisseur']
                    
                    if orientation == 'YZ':
                        # Face avant : perpendiculaire à X
                        shape = Part.makeBox(ep, L, H)
                    else:
                        # Face latérale : perpendiculaire à Y
                        shape = Part.makeBox(L, ep, H)
                    
                    pos = plinthe['position']
                    shape.translate(App.Vector(pos[0], pos[1], pos[2]))
                    
                    obj = self.gestionnaire.get_ou_creer_objet(nom_plinthe)
                    self.gestionnaire.mettre_a_jour_shape(obj, shape)
                    self.gestionnaire.appliquer_couleur(obj, plinthe['couleur'])
                    if obj not in grp_pieds.Group:
                        grp_pieds.addObject(obj)
                except Exception as e:
                    print(f"Erreur construction plinthe {plinthe['nom']}: {e}")
        
        self.gestionnaire.nettoyer_objets_orphelins(self.nom)
        
        self.doc.recompute()
        self._ajuster_vue()
        
        # Résumé
        nb_comp = len(self.compartiments)
        nb_sep = len(self.separations)
        nb_pieds = len(self.pieds)
        nb_plinthes = len(self.plinthes)
        print(f"✓ Meuble '{self.nom}' construit ({len(self.doc.Objects)} objets)")
        if nb_sep > 0:
            print(f"  {nb_comp} compartiments, {nb_sep} séparation(s)")
        if nb_pieds > 0:
            cotes_str = ", ".join(self.regles.config_plinthe.cotes)
            print(f"  {nb_pieds} pieds, {nb_plinthes} plinthe(s) ({cotes_str})")
        
        return self
    
    def _ajuster_vue(self):
        try:
            import FreeCADGui as Gui
            if Gui.ActiveDocument:
                Gui.ActiveDocument.ActiveView.fitAll()
                Gui.updateGui()
        except:
            pass
    
    # -------------------------------------------------------------------------
    # NOMENCLATURE
    # -------------------------------------------------------------------------
    
    def generer_nomenclature(self) -> str:
        lines = [
            "=" * 80,
            f"NOMENCLATURE - {self.nom}",
            "=" * 80,
            f"\nDimensions: {self.largeur} x {self.profondeur} x {self.hauteur} mm",
            f"Matériau: {self.materiau.nom} ({self.materiau.epaisseur}mm)",
        ]
        
        # Compartiments
        if len(self.compartiments) > 1:
            lines.extend(["\n" + "-" * 80, "COMPARTIMENTS", "-" * 80])
            for comp in self.compartiments:
                lines.append(f"  Compartiment {comp.index}: Y={comp.position_y:.0f}mm, Largeur={comp.largeur:.0f}mm")
                if comp.etageres:
                    lines.append(f"    Étagères: {', '.join([str(int(h)) + 'mm' for h in comp.etageres])}")
        
        # Panneaux
        lines.extend(["\n" + "-" * 80, "PANNEAUX STRUCTURE", "-" * 80])
        lines.append(f"{'Nom':<25} {'L':<8} {'l':<8} {'Ép':<6} {'m²':<8} {'Chants':<15}")
        
        total_surface = 0
        total_chant = 0
        for p in self.panneaux:
            s = p.get_surface()
            total_surface += s
            chants_str = ", ".join(p.chants.keys()) if p.chants else "-"
            chant_lin = sum(p.get_lineaire_chant().values())
            total_chant += chant_lin
            lines.append(f"{p.nom:<25} {p.longueur:<8.0f} {p.largeur:<8.0f} {p.epaisseur:<6.0f} {s:<8.4f} {chants_str:<15}")
            
            # Afficher les rainures
            if p.rainures:
                for r in p.rainures:
                    lines.append(f"    ↳ Rainure: pos={r['position']:.1f}mm, larg={r['largeur']:.1f}mm, prof={r['profondeur']:.1f}mm")
        
        # Séparations
        if self.separations:
            lines.extend(["\n" + "-" * 80, "SÉPARATIONS VERTICALES", "-" * 80])
            lines.append(f"{'Nom':<25} {'Prof':<8} {'Haut':<8} {'Ép':<6} {'m²':<8} {'Chants':<15}")
            
            for sep in self.separations:
                s = sep.get_surface()
                total_surface += s
                chants_str = ", ".join(sep.chants.keys()) if sep.chants else "-"
                chant_lin = sum(sep.get_lineaire_chant().values())
                total_chant += chant_lin
                lines.append(f"{sep.nom:<25} {sep.profondeur:<8.0f} {sep.hauteur:<8.0f} {sep.epaisseur:<6.0f} {s:<8.4f} {chants_str:<15}")
        
        lines.append(f"\nTotal panneaux + séparations: {total_surface:.4f} m²")
        lines.append(f"Total chant: {total_chant:.2f} ml")
        
        # Fond
        if self.fond:
            lines.extend(["\n" + "-" * 80, "FOND", "-" * 80])
            lines.append(self.fond.generer_info_usinage())
        
        # Portes
        if self.portes:
            lines.extend(["\n" + "-" * 80, "PORTES", "-" * 80])
            lines.append(f"  Pose: {self.regles.type_pose.value}")
            for p in self.portes:
                lines.append(f"  {p.nom}: {p.largeur:.0f} x {p.hauteur:.0f}mm, {len(p.charnieres)} charnières CLIP top")
        
        # Tiroirs
        if self.tiroirs:
            lines.extend(["\n" + "-" * 80, "TIROIRS LEGRABOX", "-" * 80])
            for t in self.tiroirs:
                lines.append(f"  {t.nom}: façade {t.largeur_facade:.0f} x {t.hauteur_facade:.0f}mm")
                lines.append(f"    Coulisse: {t.longueur_coulisse}mm, Hauteur {t.hauteur_cote.value}")
        
        # Pieds et plinthes
        if self.pieds or self.plinthes:
            lines.extend(["\n" + "-" * 80, "PIEDS ET PLINTHES", "-" * 80])
            cfg_pl = self.regles.config_plinthe
            lines.append(f"  Hauteur plinthe: {cfg_pl.hauteur:.0f}mm")
            lines.append(f"  Retrait avant: {cfg_pl.retrait:.0f}mm")
            lines.append(f"  Retrait latéral: {cfg_pl.retrait_lateral:.0f}mm")
            cotes_str = ", ".join(cfg_pl.cotes) if cfg_pl.cotes else "aucun"
            lines.append(f"  Côtés avec plinthe: {cotes_str}")
            lines.append(f"  Pieds: {len(self.pieds)} (Ø{cfg_pl.diametre_pied:.0f}mm, h={cfg_pl.hauteur_pied:.0f}mm)")
            lines.append(f"    Grille: {cfg_pl.nombre_pieds_profondeur} × {cfg_pl.nombre_pieds_largeur}")
            
            if self.plinthes:
                lines.append(f"  Assemblage angles: coupe droite, plinthe avant devant")
                lines.append(f"  Plinthes ({len(self.plinthes)}):")
                for pl in self.plinthes:
                    lines.append(f"    {pl['nom']}: {pl['longueur']:.0f} × {pl['hauteur']:.0f} × {pl['epaisseur']:.0f}mm")
        
        lines.append("\n" + "=" * 80)
        return "\n".join(lines)
    
    def exporter_nomenclature(self, chemin: str) -> str:
        """Exporte la nomenclature en fichier texte.
        
        Args:
            chemin: Chemin du fichier (ex: "/tmp/nomenclature.txt")
        
        Returns:
            Chemin du fichier créé
        """
        contenu = self.generer_nomenclature()
        with open(chemin, "w", encoding="utf-8") as f:
            f.write(contenu)
        return chemin
    
    def exporter_nomenclature_xlsx(self, chemin: str) -> str:
        """Exporte la nomenclature en fichier Excel (.xlsx).
        
        Feuilles :
          - Panneaux : liste de débit des panneaux structure + séparations
          - Portes : dimensions, charnières, type de pose
          - Tiroirs : dimensions façade, coulisse, hauteur LEGRABOX
          - Plinthes : dimensions des plinthes
          - Résumé : dimensions hors-tout, totaux surfaces et chants
        
        Args:
            chemin: Chemin du fichier (ex: "/tmp/nomenclature.xlsx")
        
        Returns:
            Chemin du fichier créé
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        
        # Styles communs
        font_titre = Font(name="Arial", bold=True, size=14)
        font_entete = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        font_normal = Font(name="Arial", size=10)
        font_total = Font(name="Arial", bold=True, size=10)
        fill_entete = PatternFill("solid", fgColor="2F5496")
        fill_total = PatternFill("solid", fgColor="D6E4F0")
        align_centre = Alignment(horizontal="center", vertical="center")
        align_droite = Alignment(horizontal="right", vertical="center")
        bordure_fine = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        def appliquer_entete(ws, row, colonnes):
            for col, texte in enumerate(colonnes, 1):
                c = ws.cell(row=row, column=col, value=texte)
                c.font = font_entete
                c.fill = fill_entete
                c.alignment = align_centre
                c.border = bordure_fine
        
        def appliquer_bordure_ligne(ws, row, nb_cols):
            for col in range(1, nb_cols + 1):
                ws.cell(row=row, column=col).border = bordure_fine
                ws.cell(row=row, column=col).font = font_normal
        
        # ==== FEUILLE 1 : PANNEAUX ====
        ws_pan = wb.active
        ws_pan.title = "Panneaux"
        
        ws_pan.cell(row=1, column=1, value=f"FICHE DE DÉBIT — {self.nom}").font = font_titre
        ws_pan.cell(row=2, column=1, value=f"Matériau : {self.materiau.nom} ({self.materiau.epaisseur}mm)").font = font_normal
        
        colonnes_pan = ["Repère", "Désignation", "L (mm)", "l (mm)", "Ép (mm)", "Qté", "Surface (m²)", "Chants", "Chant (ml)", "Rainures"]
        appliquer_entete(ws_pan, 4, colonnes_pan)
        
        row = 5
        total_surface = 0
        total_chant = 0
        repere = 1
        
        # Panneaux structure
        for p in self.panneaux:
            s = p.get_surface()
            total_surface += s
            chants_str = ", ".join(p.chants.keys()) if p.chants else "—"
            chant_lin = sum(p.get_lineaire_chant().values())
            total_chant += chant_lin
            rainures_str = f"{len(p.rainures)} fond + {len(p.rainures_cremaillere)} crém." if (p.rainures or p.rainures_cremaillere) else "—"
            
            ws_pan.cell(row=row, column=1, value=repere).alignment = align_centre
            ws_pan.cell(row=row, column=2, value=p.nom)
            ws_pan.cell(row=row, column=3, value=round(p.longueur, 1)).alignment = align_droite
            ws_pan.cell(row=row, column=4, value=round(p.largeur, 1)).alignment = align_droite
            ws_pan.cell(row=row, column=5, value=round(p.epaisseur, 1)).alignment = align_centre
            ws_pan.cell(row=row, column=6, value=1).alignment = align_centre
            ws_pan.cell(row=row, column=7, value=round(s, 4)).alignment = align_droite
            ws_pan.cell(row=row, column=8, value=chants_str)
            ws_pan.cell(row=row, column=9, value=round(chant_lin, 3)).alignment = align_droite
            ws_pan.cell(row=row, column=10, value=rainures_str)
            appliquer_bordure_ligne(ws_pan, row, len(colonnes_pan))
            repere += 1
            row += 1
        
        # Séparations
        for sep in self.separations:
            s = sep.get_surface()
            total_surface += s
            chants_str = ", ".join(sep.chants.keys()) if sep.chants else "—"
            chant_lin = sum(sep.get_lineaire_chant().values())
            total_chant += chant_lin
            
            ws_pan.cell(row=row, column=1, value=repere).alignment = align_centre
            ws_pan.cell(row=row, column=2, value=sep.nom)
            ws_pan.cell(row=row, column=3, value=round(sep.profondeur, 1)).alignment = align_droite
            ws_pan.cell(row=row, column=4, value=round(sep.hauteur, 1)).alignment = align_droite
            ws_pan.cell(row=row, column=5, value=round(sep.epaisseur, 1)).alignment = align_centre
            ws_pan.cell(row=row, column=6, value=1).alignment = align_centre
            ws_pan.cell(row=row, column=7, value=round(s, 4)).alignment = align_droite
            ws_pan.cell(row=row, column=8, value=chants_str)
            ws_pan.cell(row=row, column=9, value=round(chant_lin, 3)).alignment = align_droite
            ws_pan.cell(row=row, column=10, value="—")
            appliquer_bordure_ligne(ws_pan, row, len(colonnes_pan))
            repere += 1
            row += 1
        
        # Fond
        if self.fond:
            f = self.fond
            s_fond = (f.largeur * f.hauteur) / 1_000_000
            ws_pan.cell(row=row, column=1, value=repere).alignment = align_centre
            ws_pan.cell(row=row, column=2, value=f"Fond ({f.type_assemblage})")
            ws_pan.cell(row=row, column=3, value=round(f.largeur, 1)).alignment = align_droite
            ws_pan.cell(row=row, column=4, value=round(f.hauteur, 1)).alignment = align_droite
            ws_pan.cell(row=row, column=5, value=round(f.epaisseur, 1)).alignment = align_centre
            ws_pan.cell(row=row, column=6, value=1).alignment = align_centre
            ws_pan.cell(row=row, column=7, value=round(s_fond, 4)).alignment = align_droite
            ws_pan.cell(row=row, column=8, value="—")
            ws_pan.cell(row=row, column=9, value=0).alignment = align_droite
            ws_pan.cell(row=row, column=10, value=f.type_assemblage)
            appliquer_bordure_ligne(ws_pan, row, len(colonnes_pan))
            row += 1
        
        # Ligne totaux
        row += 1
        ws_pan.cell(row=row, column=2, value="TOTAUX").font = font_total
        ws_pan.cell(row=row, column=7, value=round(total_surface, 4)).font = font_total
        ws_pan.cell(row=row, column=7).alignment = align_droite
        ws_pan.cell(row=row, column=9, value=round(total_chant, 3)).font = font_total
        ws_pan.cell(row=row, column=9).alignment = align_droite
        for col in range(1, len(colonnes_pan) + 1):
            ws_pan.cell(row=row, column=col).fill = fill_total
            ws_pan.cell(row=row, column=col).border = bordure_fine
        
        # Largeurs colonnes
        for col, w in enumerate([6, 25, 10, 10, 8, 6, 12, 18, 10, 18], 1):
            ws_pan.column_dimensions[chr(64 + col)].width = w
        
        # ==== FEUILLE 2 : PORTES ====
        if self.portes:
            ws_p = wb.create_sheet("Portes")
            ws_p.cell(row=1, column=1, value=f"PORTES — {self.nom}").font = font_titre
            ws_p.cell(row=2, column=1, value=f"Type de pose : {self.regles.type_pose.value}").font = font_normal
            
            colonnes_p = ["Repère", "Nom", "Largeur (mm)", "Hauteur (mm)", "Ép (mm)", "Ouverture", "Nb charnières", "Chants"]
            appliquer_entete(ws_p, 4, colonnes_p)
            
            for i, p in enumerate(self.portes, 1):
                r = i + 4
                chants_str = ", ".join(p.chants.keys()) if p.chants else "—"
                ws_p.cell(row=r, column=1, value=i).alignment = align_centre
                ws_p.cell(row=r, column=2, value=p.nom)
                ws_p.cell(row=r, column=3, value=round(p.largeur, 1)).alignment = align_droite
                ws_p.cell(row=r, column=4, value=round(p.hauteur, 1)).alignment = align_droite
                ws_p.cell(row=r, column=5, value=round(p.epaisseur, 1)).alignment = align_centre
                ws_p.cell(row=r, column=6, value=p.type_ouverture.value).alignment = align_centre
                ws_p.cell(row=r, column=7, value=len(p.charnieres)).alignment = align_centre
                ws_p.cell(row=r, column=8, value=chants_str)
                appliquer_bordure_ligne(ws_p, r, len(colonnes_p))
            
            for col, w in enumerate([6, 20, 12, 12, 8, 12, 14, 18], 1):
                ws_p.column_dimensions[chr(64 + col)].width = w
        
        # ==== FEUILLE 3 : TIROIRS ====
        if self.tiroirs:
            ws_t = wb.create_sheet("Tiroirs")
            ws_t.cell(row=1, column=1, value=f"TIROIRS LEGRABOX — {self.nom}").font = font_titre
            
            colonnes_t = ["Repère", "Nom", "Façade L (mm)", "Façade H (mm)", "Ép (mm)",
                          "Coulisse (mm)", "Hauteur côté", "Larg. tiroir (mm)"]
            appliquer_entete(ws_t, 3, colonnes_t)
            
            for i, t in enumerate(self.tiroirs, 1):
                r = i + 3
                ws_t.cell(row=r, column=1, value=i).alignment = align_centre
                ws_t.cell(row=r, column=2, value=t.nom)
                ws_t.cell(row=r, column=3, value=round(t.largeur_facade, 1)).alignment = align_droite
                ws_t.cell(row=r, column=4, value=round(t.hauteur_facade, 1)).alignment = align_droite
                ws_t.cell(row=r, column=5, value=round(t.ep_facade, 1)).alignment = align_centre
                ws_t.cell(row=r, column=6, value=t.longueur_coulisse).alignment = align_centre
                ws_t.cell(row=r, column=7, value=f"{t.hauteur_cote.value} ({t.hauteur_paroi}mm)").alignment = align_centre
                ws_t.cell(row=r, column=8, value=round(t.largeur_tiroir, 1)).alignment = align_droite
                appliquer_bordure_ligne(ws_t, r, len(colonnes_t))
            
            for col, w in enumerate([6, 22, 14, 14, 8, 12, 16, 16], 1):
                ws_t.column_dimensions[chr(64 + col)].width = w
        
        # ==== FEUILLE 4 : PLINTHES ====
        if self.plinthes:
            ws_pl = wb.create_sheet("Plinthes")
            ws_pl.cell(row=1, column=1, value=f"PLINTHES ET PIEDS — {self.nom}").font = font_titre
            cfg = self.regles.config_plinthe
            ws_pl.cell(row=2, column=1, value=f"Hauteur : {cfg.hauteur:.0f}mm | Retrait av. : {cfg.retrait:.0f}mm | Retrait lat. : {cfg.retrait_lateral:.0f}mm").font = font_normal
            
            colonnes_pl = ["Repère", "Nom", "Longueur (mm)", "Hauteur (mm)", "Ép (mm)"]
            appliquer_entete(ws_pl, 4, colonnes_pl)
            
            for i, pl in enumerate(self.plinthes, 1):
                r = i + 4
                ws_pl.cell(row=r, column=1, value=i).alignment = align_centre
                ws_pl.cell(row=r, column=2, value=pl['nom'])
                ws_pl.cell(row=r, column=3, value=round(pl['longueur'], 1)).alignment = align_droite
                ws_pl.cell(row=r, column=4, value=round(pl['hauteur'], 1)).alignment = align_droite
                ws_pl.cell(row=r, column=5, value=round(pl['epaisseur'], 1)).alignment = align_droite
                appliquer_bordure_ligne(ws_pl, r, len(colonnes_pl))
            
            r += 2
            ws_pl.cell(row=r, column=1, value=f"Pieds : {len(self.pieds)} × Ø{cfg.diametre_pied:.0f}mm (h={cfg.hauteur_pied:.0f}mm)").font = font_normal
            ws_pl.cell(row=r + 1, column=1, value=f"Grille : {cfg.nombre_pieds_profondeur} × {cfg.nombre_pieds_largeur}").font = font_normal
            
            for col, w in enumerate([6, 22, 14, 14, 8], 1):
                ws_pl.column_dimensions[chr(64 + col)].width = w
        
        # ==== FEUILLE 5 : RÉSUMÉ ====
        ws_r = wb.create_sheet("Résumé")
        ws_r.cell(row=1, column=1, value=f"RÉSUMÉ — {self.nom}").font = font_titre
        
        infos = [
            ("Dimensions hors-tout", f"{self.largeur} × {self.profondeur} × {self.hauteur} mm"),
            ("Matériau", f"{self.materiau.nom} ({self.materiau.epaisseur}mm)"),
            ("Façade", f"{self.epaisseur_facade}mm"),
            ("Type de pose", self.regles.type_pose.value),
            ("", ""),
            ("Panneaux structure", str(len(self.panneaux))),
            ("Séparations", str(len(self.separations))),
            ("Compartiments", str(len(self.compartiments))),
            ("Surface panneaux", f"{total_surface:.4f} m²"),
            ("Linéaire chant", f"{total_chant:.2f} ml"),
            ("", ""),
            ("Portes", str(len(self.portes))),
            ("Tiroirs", str(len(self.tiroirs))),
            ("Pieds", str(len(self.pieds))),
            ("Plinthes", str(len(self.plinthes))),
        ]
        
        for i, (label, val) in enumerate(infos, 3):
            ws_r.cell(row=i, column=1, value=label).font = font_total if label else font_normal
            ws_r.cell(row=i, column=2, value=val).font = font_normal
            ws_r.cell(row=i, column=2).alignment = align_droite
        
        ws_r.column_dimensions['A'].width = 22
        ws_r.column_dimensions['B'].width = 35
        
        wb.save(chemin)
        return chemin


# ============================================================================
# FONCTIONS RACCOURCIS
# ============================================================================

def meuble_bas_porte(
    nom: str,
    largeur: float = 600,
    profondeur: float = 560,
    fond_rainure: bool = True,
    double: bool = True
) -> Meuble:
    """Crée un meuble bas avec porte(s)"""
    regles = ReglesAssemblage.standard()
    if fond_rainure:
        regles.config_fond = ConfigFond.standard_rainure()
    else:
        regles.config_fond = ConfigFond.standard_vissage()
    
    m = Meuble(nom, largeur, 720, profondeur, regles=regles)
    m.ajouter_cotes().ajouter_dessous(retrait=50).ajouter_dessus().ajouter_fond()
    
    if double and largeur > 450:
        m.ajouter_portes_doubles()
    else:
        m.ajouter_porte()
    
    return m


def meuble_bas_tiroirs(
    nom: str,
    largeur: float = 600,
    nombre_tiroirs: int = 3,
    hauteur_legrabox: LegraboxHauteur = LegraboxHauteur.M
) -> Meuble:
    """Crée un meuble bas avec tiroirs LEGRABOX"""
    regles = ReglesAssemblage.standard()
    regles.legrabox_hauteur = hauteur_legrabox
    
    m = Meuble(nom, largeur, 720, 560, regles=regles)
    m.ajouter_cotes().ajouter_dessous(retrait=50).ajouter_dessus().ajouter_fond()
    m.ajouter_tiroirs_legrabox(nombre_tiroirs)
    
    return m


# ============================================================================
# EXEMPLE
# ============================================================================

if __name__ == "__main__" or True:
    
    # ========================================
    # CONFIGURATION
    # ========================================
    
    LARGEUR = 1200              # Largeur totale du meuble
    HAUTEUR = 720
    PROFONDEUR = 560
    
    # Type de façade: "porte" ou "tiroirs" ou "sans"
    TYPE_FACADE = "porte"
    TYPE_POSE = TypePose.EN_APPLIQUE  # EN_APPLIQUE / SEMI_APPLIQUE / ENCLOISONNEE
    
    # Configuration du fond
    TYPE_FOND = TypeAssemblageFond.RAINURE
    FOND_EPAISSEUR = 3
    FOND_DECALAGE = 0
    FOND_PROFONDEUR_RAINURE = 8
    FOND_JEU_RAINURE = 1
    FOND_DISTANCE_CHANT = 10
    FOND_HAUTEUR_IMPOSEE = None
    FOND_POSITION_VERTICALE = "bas"
    
    # Configuration des séparations
    NOMBRE_COMPARTIMENTS = 3    # 1 = pas de séparation, 2+ = avec séparations
    LARGEURS_COMPARTIMENTS = None  # ou [300, 400, 462] pour personnaliser
    RETRAIT_ARRIERE = 0
    RETRAIT_AVANT = 0
    
    # Configuration des crémaillères
    CREMAILLERE_LARGEUR = 16        # Largeur du profilé aluminium
    CREMAILLERE_PROFONDEUR = 7      # Profondeur d'encastrement
    CREMAILLERE_DISTANCE_AVANT = 37 # Distance depuis l'avant
    CREMAILLERE_DISTANCE_ARRIERE = 37
    CREMAILLERE_DOUBLE = True       # 2 crémaillères par panneau
    
    # Configuration des étagères
    ETAGERE_JEU_LATERAL = 1         # Jeu de chaque côté
    ETAGERE_JEU_ARRIERE = 5         # Jeu avec le fond
    ETAGERE_RETRAIT_AVANT = 20      # Retrait depuis l'avant
    
    # Configuration des pieds et plinthes
    PLINTHE_COTES = ["avant", "gauche", "droite"]  # ou ["avant"] ou ["avant", "gauche"]
    PLINTHE_EPAISSEUR = 16          # Épaisseur de la plinthe
    PLINTHE_RETRAIT = 30            # Retrait par rapport à la face avant (façade)
    PLINTHE_RETRAIT_LATERAL = 16    # Retrait des plinthes latérales depuis le côté
    PIEDS_PROFONDEUR = 2            # Nombre de pieds en profondeur
    PIEDS_LARGEUR = 3               # Nombre de pieds en largeur
    
    # ========================================
    
    # Créer la configuration du fond
    config_fond = ConfigFond(
        type_assemblage=TYPE_FOND,
        epaisseur=FOND_EPAISSEUR,
        decalage_arriere=FOND_DECALAGE,
        profondeur_rainure=FOND_PROFONDEUR_RAINURE,
        jeu_fond_rainure=FOND_JEU_RAINURE,
        distance_rainure_chant=FOND_DISTANCE_CHANT,
        hauteur_imposee=FOND_HAUTEUR_IMPOSEE,
        position_verticale=FOND_POSITION_VERTICALE
    )
    
    # Créer la configuration des séparations
    config_sep = ConfigSeparations(
        nombre_compartiments=NOMBRE_COMPARTIMENTS,
        largeurs_compartiments=LARGEURS_COMPARTIMENTS,
        epaisseur=19,
        retrait_arriere=RETRAIT_ARRIERE,
        retrait_avant=RETRAIT_AVANT,
        chant_avant=True,
        epaisseur_chant=1
    )
    
    # Créer la configuration des crémaillères
    config_crem = ConfigCremaillere(
        largeur=CREMAILLERE_LARGEUR,
        profondeur=CREMAILLERE_PROFONDEUR,
        distance_avant=CREMAILLERE_DISTANCE_AVANT,
        distance_arriere=CREMAILLERE_DISTANCE_ARRIERE,
        double=CREMAILLERE_DOUBLE
    )
    
    # Créer la configuration des étagères
    config_etag = ConfigEtagere(
        jeu_lateral=ETAGERE_JEU_LATERAL,
        jeu_arriere=ETAGERE_JEU_ARRIERE,
        retrait_avant=ETAGERE_RETRAIT_AVANT,
        chant_avant=True
    )
    
    # Créer la configuration des pieds et plinthes
    config_plinthe = ConfigPlinthe(
        epaisseur=PLINTHE_EPAISSEUR,
        retrait=PLINTHE_RETRAIT,
        retrait_lateral=PLINTHE_RETRAIT_LATERAL,
        cotes=PLINTHE_COTES,
        nombre_pieds_profondeur=PIEDS_PROFONDEUR,
        nombre_pieds_largeur=PIEDS_LARGEUR
    )
    
    # Créer les règles
    regles = ReglesAssemblage.standard()
    regles.configurer_fond(config_fond)
    regles.configurer_separations(config_sep)
    regles.configurer_cremailleres(config_crem)
    regles.configurer_etageres(config_etag)
    regles.configurer_plinthe(config_plinthe)
    regles.configurer_pose(TYPE_POSE)
    
    # Créer le meuble
    meuble = Meuble("MonMeuble", LARGEUR, HAUTEUR, PROFONDEUR, regles=regles)
    meuble.ajouter_cotes()
    meuble.ajouter_dessous(retrait=50)
    meuble.ajouter_dessus(type="traverses")
    meuble.ajouter_fond()
    meuble.ajouter_separations()
    meuble.ajouter_cremailleres()
    meuble.ajouter_pieds_et_plinthes()  # Pieds + plinthes sur 3 côtés
    
    # Ajouter des étagères dans les compartiments
    if NOMBRE_COMPARTIMENTS >= 2:
        meuble.ajouter_etagere_compartiment(0, hauteur_depuis_bas=200)
        meuble.ajouter_etagere_compartiment(0, hauteur_depuis_bas=400)
        meuble.ajouter_etagere_compartiment(1, hauteur_depuis_bas=300)
        meuble.ajouter_etagere_compartiment(2, hauteur_depuis_bas=250)
    
    # Façades
    if TYPE_FACADE == "tiroirs":
        meuble.ajouter_tiroirs_legrabox(3)
    elif TYPE_FACADE == "porte":
        if meuble.compartiments and len(meuble.compartiments) > 1:
            # Compartiment 0 : 2 tiroirs F + 1 tiroir K
            meuble.ajouter_facades_compartiment(0, [
                ("tiroir", LegraboxHauteur.F),
                ("tiroir", LegraboxHauteur.F),
                ("tiroir", LegraboxHauteur.K),
            ])
            # Compartiment 1 : 1 tiroir K en bas + 1 porte en haut
            meuble.ajouter_facades_compartiment(1, [
                ("tiroir", LegraboxHauteur.K),
                ("porte", TypeOuverture.DROITE),
            ])
            # Compartiment 2 : 1 porte pleine
            meuble.ajouter_portes_compartiments([
                (2, TypeOuverture.GAUCHE),
            ])
        elif LARGEUR > 450:
            meuble.ajouter_portes_doubles()
        else:
            meuble.ajouter_porte()
    
    meuble.construire()
    print(meuble.generer_nomenclature())
    
    # Export nomenclature
    import os
    dossier = os.path.dirname(App.ActiveDocument.FileName) if App.ActiveDocument and App.ActiveDocument.FileName else "/tmp"
    meuble.exporter_nomenclature(os.path.join(dossier, f"{meuble.nom}_nomenclature.txt"))
    meuble.exporter_nomenclature_xlsx(os.path.join(dossier, f"{meuble.nom}_nomenclature.xlsx"))
    print(f"Nomenclature exportée dans {dossier}")
